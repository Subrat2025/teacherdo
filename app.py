"""
Teacher Verification Website — Flask Backend v4
Chikiti Block, Ganjam District | 2026
New: Bulk Excel import into any proforma
"""
from flask import (Flask, render_template_string, request, jsonify,
                   session, redirect, url_for, send_file)
import json, os, datetime, io, re
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "chikiti_block_2026_secure_key_xyz"

ADMIN_USER = "Subrat"
ADMIN_PASS = "Subrat@888"

# ══ Embedded HTML Templates ══

HOME_HTML = '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>Teacher Proforma Verification — Chikiti Block</title>\n<style>\n  :root{--navy:#0d2366;--steel:#1f4e79;--sky:#dce6f1;--green:#2e7d32;--amber:#f57f17;--red:#c62828;--white:#fff;--grey:#f5f7fa}\n  *{box-sizing:border-box;margin:0;padding:0}\n  body{font-family:\'Segoe UI\',Arial,sans-serif;background:var(--grey);min-height:100vh;display:flex;flex-direction:column}\n  /* ── Header ── */\n  header{background:linear-gradient(135deg,var(--navy) 0%,var(--steel) 100%);color:var(--white);padding:0}\n  .header-inner{max-width:1000px;margin:auto;padding:18px 24px;display:flex;align-items:center;gap:18px}\n  .header-logo{width:58px;height:58px;background:var(--white);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:28px;flex-shrink:0}\n  .header-text h1{font-size:1.35rem;font-weight:700;letter-spacing:.3px}\n  .header-text p{font-size:.82rem;opacity:.85;margin-top:3px}\n  /* ── Hero ── */\n  .hero{background:linear-gradient(135deg,#e8f0fe 0%,#d0e8ff 100%);padding:52px 24px 44px;text-align:center}\n  .hero h2{font-size:1.9rem;color:var(--navy);font-weight:700;margin-bottom:10px}\n  .hero p{color:#444;font-size:1rem;max-width:620px;margin:auto;line-height:1.6}\n  /* ── Cards ── */\n  .cards{max-width:800px;margin:-28px auto 0;padding:0 24px 60px;display:grid;grid-template-columns:1fr 1fr;gap:24px}\n  @media(max-width:600px){.cards{grid-template-columns:1fr}}\n  .card{background:var(--white);border-radius:14px;box-shadow:0 4px 20px rgba(0,0,0,.1);padding:36px 28px;text-align:center;transition:transform .2s,box-shadow .2s;text-decoration:none;color:inherit;display:block}\n  .card:hover{transform:translateY(-4px);box-shadow:0 8px 32px rgba(0,0,0,.15)}\n  .card-icon{font-size:3rem;margin-bottom:14px}\n  .card h3{font-size:1.2rem;color:var(--navy);font-weight:700;margin-bottom:8px}\n  .card p{font-size:.88rem;color:#555;line-height:1.5;margin-bottom:20px}\n  .btn{display:inline-block;padding:11px 28px;border-radius:8px;font-size:.92rem;font-weight:600;text-decoration:none;transition:opacity .2s}\n  .btn-teacher{background:var(--navy);color:var(--white)}\n  .btn-admin{background:var(--steel);color:var(--white)}\n  .btn:hover{opacity:.88}\n  /* ── Info bar ── */\n  .info-bar{background:var(--navy);color:var(--white);text-align:center;padding:14px 24px;font-size:.84rem;opacity:.92}\n  .info-bar span{margin:0 14px}\n  /* ── Footer ── */\n  footer{margin-top:auto;background:#1a1a2e;color:#aaa;text-align:center;padding:16px;font-size:.8rem}\n</style>\n</head>\n<body>\n\n<header>\n  <div class="header-inner">\n    <div class="header-logo">🏫</div>\n    <div class="header-text">\n      <h1>Teacher Proforma Verification System</h1>\n      <p>Chikiti Block, Ganjam District, Odisha &nbsp;|&nbsp; 2026</p>\n    </div>\n  </div>\n</header>\n\n<div class="hero">\n  <h2>Welcome to the Verification Portal</h2>\n  <p>Teachers can log in to view, verify and update their proforma data. Administrators can manage all records and export updated data to Excel.</p>\n</div>\n\n<div class="cards">\n  <a class="card" href="/teacher/login">\n    <div class="card-icon">👨\u200d🏫</div>\n    <h3>Teacher Login</h3>\n    <p>Select your school and enter your date of birth to access and verify your proforma data.</p>\n    <span class="btn btn-teacher">Teacher Login →</span>\n  </a>\n\n  <a class="card" href="/admin/login">\n    <div class="card-icon">🛡️</div>\n    <h3>Admin Login</h3>\n    <p>Authorised administrators can manage all teacher records, view updates and export data.</p>\n    <span class="btn btn-admin">Admin Login →</span>\n  </a>\n</div>\n\n<div class="info-bar">\n  <span>📊 Total Teachers: 43</span>\n  <span>🏫 Schools: 12</span>\n  <span>📋 Proformas I – VIII</span>\n  <span>🔒 Secure &amp; Encrypted</span>\n</div>\n\n<footer>\n  &copy; 2026 Chikiti Block Education Office, Ganjam District &nbsp;|&nbsp; Teacher Proforma Verification System\n</footer>\n\n</body>\n</html>\n'

TEACHER_LOGIN_HTML = '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>Teacher Login — Chikiti Block</title>\n<style>\n  :root{--navy:#0d2366;--steel:#1f4e79;--sky:#dce6f1;--green:#2e7d32;--amber:#f57f17;--red:#c62828}\n  *{box-sizing:border-box;margin:0;padding:0}\n  body{font-family:\'Segoe UI\',Arial,sans-serif;background:linear-gradient(135deg,#e8f0fe 0%,#c8deff 100%);min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px}\n  .back-link{position:fixed;top:16px;left:16px;color:var(--navy);text-decoration:none;font-size:.9rem;font-weight:600;background:rgba(255,255,255,.85);padding:8px 14px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.1)}\n  .back-link:hover{background:var(--white)}\n  /* Card */\n  .card{background:#fff;border-radius:18px;box-shadow:0 8px 40px rgba(13,35,102,.15);padding:44px 40px 40px;max-width:480px;width:100%}\n  .logo{text-align:center;margin-bottom:28px}\n  .logo .icon{font-size:3.2rem;margin-bottom:8px}\n  .logo h1{font-size:1.5rem;color:var(--navy);font-weight:700}\n  .logo p{font-size:.85rem;color:#666;margin-top:4px}\n  /* Form */\n  .field{margin-bottom:20px}\n  label{display:block;font-size:.88rem;font-weight:600;color:var(--steel);margin-bottom:6px}\n  select,input[type=text],input[type=date]{\n    width:100%;padding:11px 14px;border:2px solid #d0daf0;border-radius:9px;\n    font-size:.95rem;color:#222;background:#fdfdff;transition:border-color .2s;outline:none}\n  select:focus,input:focus{border-color:var(--navy)}\n  .hint{font-size:.78rem;color:#888;margin-top:5px}\n  .btn-login{\n    width:100%;padding:14px;background:linear-gradient(135deg,var(--navy),var(--steel));\n    color:#fff;border:none;border-radius:10px;font-size:1rem;font-weight:700;\n    cursor:pointer;transition:opacity .2s;margin-top:8px}\n  .btn-login:hover{opacity:.9}\n  .btn-login:disabled{opacity:.6;cursor:not-allowed}\n  /* Error / info */\n  .alert{padding:12px 16px;border-radius:9px;font-size:.88rem;margin-bottom:16px;display:none}\n  .alert.error{background:#ffeaea;color:var(--red);border:1px solid #ffcdd2}\n  .alert.success{background:#e8f5e9;color:var(--green);border:1px solid #c8e6c9}\n  .alert.show{display:block}\n  /* Spinner */\n  .spinner{display:none;width:20px;height:20px;border:3px solid rgba(255,255,255,.4);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;margin:auto}\n  @keyframes spin{to{transform:rotate(360deg)}}\n  /* Info box */\n  .info-box{background:#f0f4ff;border-left:4px solid var(--navy);border-radius:0 8px 8px 0;padding:12px 14px;margin-bottom:22px;font-size:.84rem;color:var(--navy);line-height:1.5}\n</style>\n</head>\n<body>\n\n<a class="back-link" href="/">← Home</a>\n\n<div class="card">\n  <div class="logo">\n    <div class="icon">👨\u200d🏫</div>\n    <h1>Teacher Login</h1>\n    <p>Chikiti Block, Ganjam District — 2026</p>\n  </div>\n\n  <div class="info-box">\n    🔐 Select your school and enter your date of birth (as recorded in the office). Your proforma data will be displayed for verification.\n  </div>\n\n  <div class="alert error" id="alertBox"></div>\n\n  <div class="field">\n    <label for="schoolSelect">🏫 Select Your School *</label>\n    <select id="schoolSelect">\n      <option value="">— Choose your school —</option>\n      {% for school in schools %}\n      <option value="{{ school }}">{{ school }}</option>\n      {% endfor %}\n    </select>\n  </div>\n\n  <div class="field">\n    <label for="dobInput">📅 Date of Birth *</label>\n    <input type="date" id="dobInput" placeholder="DD/MM/YYYY">\n    <div class="hint">Enter your date of birth as registered in the proforma.</div>\n  </div>\n\n  <button class="btn-login" id="loginBtn" onclick="doLogin()">\n    <span id="btnText">Login &amp; View My Data →</span>\n    <div class="spinner" id="spinner"></div>\n  </button>\n</div>\n\n<script>\nasync function doLogin(){\n  const school = document.getElementById(\'schoolSelect\').value.trim();\n  const dobRaw  = document.getElementById(\'dobInput\').value.trim();\n  const alert   = document.getElementById(\'alertBox\');\n  alert.className = \'alert\';\n\n  if(!school){ showErr(\'Please select your school.\'); return; }\n  if(!dobRaw){  showErr(\'Please enter your date of birth.\'); return; }\n\n  // Convert YYYY-MM-DD (from date input) → DD/MM/YYYY\n  const parts = dobRaw.split(\'-\');\n  const dob   = parts.length===3 ? `${parts[2]}/${parts[1]}/${parts[0]}` : dobRaw;\n\n  setBusy(true);\n  try{\n    const res  = await fetch(\'/teacher/login\',{\n      method:\'POST\',\n      headers:{\'Content-Type\':\'application/json\'},\n      body: JSON.stringify({school, dob})\n    });\n    const data = await res.json();\n    if(data.success){\n      window.location.href = \'/teacher/dashboard\';\n    } else {\n      showErr(data.error || \'Login failed. Please check your details.\');\n    }\n  } catch(e){ showErr(\'Network error. Please try again.\'); }\n  finally{ setBusy(false); }\n}\n\nfunction showErr(msg){\n  const a = document.getElementById(\'alertBox\');\n  a.textContent = msg;\n  a.className = \'alert error show\';\n}\nfunction setBusy(on){\n  document.getElementById(\'loginBtn\').disabled = on;\n  document.getElementById(\'btnText\').style.display = on?\'none\':\'block\';\n  document.getElementById(\'spinner\').style.display = on?\'block\':\'none\';\n}\n\n// Allow Enter key\ndocument.addEventListener(\'keydown\', e=>{ if(e.key===\'Enter\') doLogin(); });\n</script>\n</body>\n</html>\n'

TEACHER_DASHBOARD_HTML = '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>My Proforma Data — Teacher Dashboard</title>\n<style>\n  :root{--navy:#0d2366;--steel:#1f4e79;--sky:#dce6f1;--green:#2e7d32;--amber:#e65100;--red:#c62828;--grey:#f5f7fa}\n  *{box-sizing:border-box;margin:0;padding:0}\n  body{font-family:\'Segoe UI\',Arial,sans-serif;background:var(--grey);min-height:100vh}\n\n  /* ── Top bar ── */\n  .topbar{background:linear-gradient(135deg,var(--navy),var(--steel));color:#fff;padding:14px 24px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px}\n  .topbar-left{display:flex;align-items:center;gap:12px}\n  .topbar-left .icon{font-size:1.8rem}\n  .topbar-left h1{font-size:1.1rem;font-weight:700}\n  .topbar-left p{font-size:.78rem;opacity:.8}\n  .topbar-right{display:flex;gap:10px;align-items:center}\n  .school-badge{background:rgba(255,255,255,.15);padding:7px 14px;border-radius:20px;font-size:.82rem}\n  .btn-logout{background:rgba(255,255,255,.12);color:#fff;border:1px solid rgba(255,255,255,.3);padding:8px 16px;border-radius:8px;cursor:pointer;font-size:.85rem;text-decoration:none;transition:background .2s}\n  .btn-logout:hover{background:rgba(255,255,255,.25)}\n\n  /* ── Container ── */\n  .container{max-width:1100px;margin:0 auto;padding:24px 18px 60px}\n\n  /* ── Progress bar ── */\n  .progress-wrap{background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.07);padding:18px 22px;margin-bottom:22px}\n  .progress-label{font-size:.85rem;color:#555;margin-bottom:6px}\n  .progress-bar{height:9px;background:#e0e0e0;border-radius:5px;overflow:hidden}\n  .progress-fill{height:100%;background:linear-gradient(90deg,#2196f3,#0d2366);border-radius:5px;transition:width .6s}\n  .progress-stats{display:flex;gap:18px;margin-top:12px;flex-wrap:wrap}\n  .stat-item{font-size:.82rem;color:#555}\n  .stat-item strong{color:var(--navy)}\n\n  /* ── Teacher card ── */\n  .teacher-card{background:#fff;border-radius:14px;box-shadow:0 3px 18px rgba(0,0,0,.09);margin-bottom:28px;overflow:hidden}\n  .teacher-card-header{background:linear-gradient(135deg,var(--navy),var(--steel));color:#fff;padding:18px 24px;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap}\n  .teacher-name{font-size:1.15rem;font-weight:700}\n  .teacher-meta{font-size:.82rem;opacity:.85;margin-top:3px}\n  .status-badge{padding:6px 16px;border-radius:20px;font-size:.82rem;font-weight:700;text-transform:uppercase}\n  .badge-PENDING{background:#fff3e0;color:#e65100;border:1.5px solid #ffcc02}\n  .badge-VERIFIED{background:#e8f5e9;color:#1b5e20;border:1.5px solid #66bb6a}\n  .badge-UPDATED{background:#e3f2fd;color:#0d47a1;border:1.5px solid #42a5f5}\n\n  /* ── Field groups ── */\n  .teacher-body{padding:22px 24px}\n  .field-group{margin-bottom:22px}\n  .group-title{font-size:.92rem;font-weight:700;color:var(--steel);margin-bottom:12px;padding-bottom:6px;border-bottom:2px solid var(--sky);display:flex;align-items:center;gap:6px}\n  .fields-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:12px}\n  .field-item{border:1.5px solid #e8edf5;border-radius:9px;padding:10px 13px;background:#fdfdff;transition:border-color .2s}\n  .field-item.changed{border-color:#1976d2;background:#f0f6ff}\n  .field-label{font-size:.75rem;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.4px;margin-bottom:4px}\n  .field-value{font-size:.92rem;color:#222;word-break:break-word}\n  .field-input{width:100%;border:none;background:transparent;font-size:.92rem;color:#1a237e;outline:none;font-family:inherit;padding:0}\n  .field-input::placeholder{color:#bbb}\n\n  /* ── Action buttons ── */\n  .action-bar{padding:16px 24px 22px;background:#f8f9ff;border-top:1px solid #e8edf5;display:flex;gap:12px;flex-wrap:wrap;align-items:center}\n  .btn{padding:10px 22px;border-radius:9px;font-size:.9rem;font-weight:600;border:none;cursor:pointer;transition:all .2s;display:flex;align-items:center;gap:7px}\n  .btn-verify{background:linear-gradient(135deg,#2e7d32,#388e3c);color:#fff}\n  .btn-verify:hover{opacity:.9;transform:translateY(-1px)}\n  .btn-save{background:linear-gradient(135deg,#1565c0,#1976d2);color:#fff}\n  .btn-save:hover{opacity:.9;transform:translateY(-1px)}\n  .btn-edit{background:#fff;color:var(--navy);border:2px solid var(--navy)}\n  .btn-edit:hover{background:var(--sky)}\n  .btn-cancel{background:#fff;color:#666;border:2px solid #ddd}\n  .btn-cancel:hover{background:#f5f5f5}\n  .btn:disabled{opacity:.55;cursor:not-allowed;transform:none}\n  .changes-count{font-size:.82rem;color:#1565c0;background:#e3f2fd;padding:5px 12px;border-radius:20px;margin-left:auto}\n\n  /* ── Success banner ── */\n  .success-banner{background:#e8f5e9;border:1.5px solid #81c784;color:#1b5e20;border-radius:10px;padding:13px 18px;margin-bottom:18px;display:flex;align-items:center;gap:10px;display:none}\n  .success-banner.show{display:flex}\n\n  /* ── Toast ── */\n  .toast{position:fixed;bottom:28px;right:28px;padding:14px 22px;border-radius:10px;color:#fff;font-size:.9rem;font-weight:600;z-index:9999;opacity:0;transform:translateY(20px);transition:all .35s;pointer-events:none;max-width:340px;box-shadow:0 6px 24px rgba(0,0,0,.2)}\n  .toast.show{opacity:1;transform:translateY(0)}\n  .toast.success{background:#2e7d32}\n  .toast.error{background:#c62828}\n  .toast.info{background:#0d2366}\n\n  /* ── Spinner ── */\n  .spinner{width:16px;height:16px;border:2.5px solid rgba(255,255,255,.4);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite}\n  @keyframes spin{to{transform:rotate(360deg)}}\n\n  /* ── Loading overlay ── */\n  #loadingOverlay{position:fixed;inset:0;background:rgba(255,255,255,.85);display:flex;align-items:center;justify-content:center;z-index:9000;flex-direction:column;gap:14px}\n  #loadingOverlay .sp{width:44px;height:44px;border:5px solid #dce6f1;border-top-color:var(--navy);border-radius:50%;animation:spin 1s linear infinite}\n  #loadingOverlay p{color:var(--navy);font-weight:600}\n\n  /* ── Edit mode indicator ── */\n  .edit-mode-bar{background:#fff3cd;border:1px solid #ffc107;color:#856404;padding:10px 16px;border-radius:8px;font-size:.85rem;margin-bottom:14px;display:none;align-items:center;gap:8px}\n  .edit-mode-bar.show{display:flex}\n\n  @media(max-width:600px){\n    .topbar{flex-direction:column;align-items:flex-start}\n    .fields-grid{grid-template-columns:1fr}\n    .action-bar{justify-content:center}\n  }\n</style>\n</head>\n<body>\n\n<!-- Loading Overlay -->\n<div id="loadingOverlay">\n  <div class="sp"></div>\n  <p>Loading your proforma data…</p>\n</div>\n\n<!-- Top Bar -->\n<div class="topbar">\n  <div class="topbar-left">\n    <div class="icon">👨\u200d🏫</div>\n    <div>\n      <h1>My Proforma Data</h1>\n      <p>Chikiti Block Verification Portal · 2026</p>\n    </div>\n  </div>\n  <div class="topbar-right">\n    <div class="school-badge">🏫 {{ school }}</div>\n    <a href="/teacher/logout" class="btn-logout">🚪 Logout</a>\n  </div>\n</div>\n\n<!-- Main container -->\n<div class="container">\n\n  <!-- Progress bar -->\n  <div class="progress-wrap" id="progressWrap" style="display:none">\n    <div class="progress-label">Verification Progress</div>\n    <div class="progress-bar"><div class="progress-fill" id="progressFill" style="width:0%"></div></div>\n    <div class="progress-stats" id="progressStats"></div>\n  </div>\n\n  <!-- Teacher records inserted here -->\n  <div id="teacherRecords"></div>\n\n  <!-- No data message -->\n  <div id="noData" style="display:none;background:#fff;border-radius:14px;padding:40px;text-align:center;color:#999;font-size:1.1rem">\n    ⚠️ No records found. Please <a href="/teacher/logout">logout</a> and try again.\n  </div>\n\n</div>\n\n<!-- Toast -->\n<div class="toast" id="toast"></div>\n\n<script>\n/* ── State ── */\nlet allTeachers = [], fieldLabels = {}, fieldGroups = [];\nlet editModes = {}; // { "NAME||PROFORMA": bool }\n\n/* ── Boot ── */\nwindow.onload = async () => {\n  try {\n    const res  = await fetch(\'/api/teacher/mydata\');\n    const data = await res.json();\n    allTeachers = data.teachers || [];\n    fieldLabels = data.field_labels || {};\n    fieldGroups = data.field_groups || [];\n    renderAll();\n  } catch(e) {\n    document.getElementById(\'noData\').style.display = \'block\';\n  } finally {\n    document.getElementById(\'loadingOverlay\').style.display = \'none\';\n  }\n};\n\n/* ── Render all teachers ── */\nfunction renderAll(){\n  if(!allTeachers.length){ document.getElementById(\'noData\').style.display=\'block\'; return; }\n\n  // Progress\n  const total = allTeachers.length;\n  const done  = allTeachers.filter(t=>t.status===\'VERIFIED\'||t.status===\'UPDATED\').length;\n  if(total > 1){\n    document.getElementById(\'progressWrap\').style.display = \'block\';\n    document.getElementById(\'progressFill\').style.width = `${Math.round(done/total*100)}%`;\n    document.getElementById(\'progressStats\').innerHTML =\n      allTeachers.map(t=>`<div class="stat-item"><strong>${t.teacher_name}</strong>: \n        <span style="color:${statusColor(t.status)}">${t.status}</span></div>`).join(\'\');\n  }\n\n  const container = document.getElementById(\'teacherRecords\');\n  container.innerHTML = \'\';\n  allTeachers.forEach((t,i)=>{\n    container.insertAdjacentHTML(\'beforeend\', buildCard(t, i));\n  });\n}\n\nfunction statusColor(s){ return s===\'VERIFIED\'?\'#2e7d32\':s===\'UPDATED\'?\'#1565c0\':\'#e65100\'; }\n\n/* ── Build one teacher card ── */\nfunction buildCard(t, idx){\n  const key  = cardKey(t);\n  const isEdit = editModes[key] || false;\n  const pf   = t.proforma || \'\';\n  const ts   = t.last_updated ? ` · Last updated: ${t.last_updated}` : \'\';\n  const chCount = (t.change_history||[]).length;\n\n  let groupsHtml = \'\';\n  for(const grp of fieldGroups){\n    let fieldsHtml = \'\';\n    for(const fk of grp.fields){\n      if(fk===\'teacher_name\'||fk===\'sl_no\') continue;\n      const label = fieldLabels[fk] || fk;\n      const val   = t[fk] || \'\';\n      const isDate = fk.startsWith(\'date_\')||fk===\'dob\';\n      fieldsHtml += `\n        <div class="field-item" id="fi_${idx}_${fk}">\n          <div class="field-label">${label}</div>\n          ${isEdit ? `<input class="field-input" id="inp_${idx}_${fk}" \n            type="${isDate?\'text\':\'text\'}" value="${escHtml(val)}" \n            placeholder="${isDate?\'DD/MM/YYYY\':\'\'}"\n            oninput="markChanged(${idx},\'${fk}\')">` \n            : `<div class="field-value">${escHtml(val)||\'<span style="color:#bbb">—</span>\'}</div>`}\n        </div>`;\n    }\n    groupsHtml += `\n      <div class="field-group">\n        <div class="group-title">${grp.title}</div>\n        <div class="fields-grid">${fieldsHtml}</div>\n      </div>`;\n  }\n\n  // Determine action bar content based on status & edit mode\n  let actionHtml = \'\';\n  if(t.status === \'VERIFIED\'){\n    actionHtml = `<span style="color:#2e7d32;font-weight:600">✅ This record has been verified.</span>\n      <button class="btn btn-edit" onclick="setEdit(\'${key}\',${idx},true)">✏️ Edit &amp; Correct</button>`;\n  } else if(isEdit){\n    actionHtml = `\n      <button class="btn btn-save" id="saveBtn_${idx}" onclick="saveUpdates(${idx})" disabled>\n        💾 Save Changes\n      </button>\n      <button class="btn btn-verify" onclick="verifyRecord(${idx})">✅ Verify as Correct</button>\n      <button class="btn btn-cancel" onclick="setEdit(\'${key}\',${idx},false)">✖ Cancel</button>\n      <span class="changes-count" id="chgCount_${idx}" style="display:none">0 changes</span>`;\n  } else {\n    actionHtml = `\n      <button class="btn btn-verify" onclick="verifyRecord(${idx})">✅ All data is correct — Verify</button>\n      <button class="btn btn-edit" onclick="setEdit(\'${key}\',${idx},true)">✏️ Edit / Correct Data</button>`;\n  }\n\n  const changesBadge = chCount>0 ? `<span style="font-size:.78rem;background:rgba(255,255,255,.2);padding:3px 10px;border-radius:12px;margin-left:8px">${chCount} edit(s)</span>` : \'\';\n\n  return `\n    <div class="teacher-card" id="card_${idx}">\n      <div class="teacher-card-header">\n        <div>\n          <div class="teacher-name">👤 ${escHtml(t.teacher_name)} ${changesBadge}</div>\n          <div class="teacher-meta">${pf} &nbsp;·&nbsp; S.No. ${t.sl_no||\'—\'}${ts}</div>\n        </div>\n        <div class="status-badge badge-${t.status||\'PENDING\'}">${t.status||\'PENDING\'}</div>\n      </div>\n      ${isEdit?`<div class="edit-mode-bar show">✏️ Edit mode active — modify the fields below then click <strong>Save Changes</strong> or <strong>Verify</strong>.</div>`:\'\'}\n      <div class="teacher-body" id="body_${idx}">${groupsHtml}</div>\n      <div class="action-bar">${actionHtml}</div>\n    </div>`;\n}\n\nfunction cardKey(t){ return t.teacher_name+\'||\'+t.proforma; }\n\n/* ── Edit mode toggle ── */\nfunction setEdit(key, idx, on){\n  editModes[key] = on;\n  // Re-render card in place\n  const card   = document.getElementById(`card_${idx}`);\n  const t      = allTeachers[idx];\n  card.outerHTML = buildCard(t, idx);\n}\n\n/* ── Track changes ── */\nfunction markChanged(idx, fk){\n  const t   = allTeachers[idx];\n  const inp = document.getElementById(`inp_${idx}_${fk}`);\n  const fi  = document.getElementById(`fi_${idx}_${fk}`);\n  if(!inp||!fi) return;\n  const changed = inp.value.trim() !== (t[fk]||\'\').trim();\n  fi.classList.toggle(\'changed\', changed);\n  updateChangeCount(idx);\n}\n\nfunction updateChangeCount(idx){\n  const t  = allTeachers[idx];\n  let cnt  = 0;\n  for(const fk of Object.keys(fieldLabels)){\n    const inp = document.getElementById(`inp_${idx}_${fk}`);\n    if(inp && inp.value.trim() !== (t[fk]||\'\').trim()) cnt++;\n  }\n  const cc = document.getElementById(`chgCount_${idx}`);\n  const sb = document.getElementById(`saveBtn_${idx}`);\n  if(cc){ cc.textContent=`${cnt} change(s)`; cc.style.display=cnt?\'inline-block\':\'none\'; }\n  if(sb) sb.disabled = (cnt===0);\n}\n\n/* ── Collect edits ── */\nfunction collectEdits(idx){\n  const t = allTeachers[idx];\n  const updated = {};\n  for(const fk of Object.keys(fieldLabels)){\n    const inp = document.getElementById(`inp_${idx}_${fk}`);\n    if(inp && inp.value.trim() !== (t[fk]||\'\').trim()){\n      updated[fk] = inp.value.trim();\n    }\n  }\n  return updated;\n}\n\n/* ── Save updates ── */\nasync function saveUpdates(idx){\n  const t       = allTeachers[idx];\n  const fields  = collectEdits(idx);\n  if(!Object.keys(fields).length){ showToast(\'No changes to save.\',\'info\'); return; }\n\n  const sb = document.getElementById(`saveBtn_${idx}`);\n  if(sb){ sb.disabled=true; sb.innerHTML=\'<div class="spinner"></div> Saving…\'; }\n\n  try{\n    const res  = await fetch(\'/api/teacher/update\',{\n      method:\'POST\',\n      headers:{\'Content-Type\':\'application/json\'},\n      body: JSON.stringify({teacher_name:t.teacher_name, proforma:t.proforma, fields})\n    });\n    const data = await res.json();\n    if(data.success){\n      // Apply changes locally\n      Object.assign(t, fields);\n      t.status = \'UPDATED\';\n      t.last_updated = data.timestamp || \'\';\n      editModes[cardKey(t)] = false;\n      const card = document.getElementById(`card_${idx}`);\n      card.outerHTML = buildCard(t, idx);\n      showToast(\'✅ \'+data.message, \'success\');\n    } else {\n      showToast(\'❌ \'+(data.error||\'Update failed.\'), \'error\');\n      if(sb){ sb.disabled=false; sb.innerHTML=\'💾 Save Changes\'; }\n    }\n  } catch(e){\n    showToast(\'❌ Network error. Please retry.\', \'error\');\n    if(sb){ sb.disabled=false; sb.innerHTML=\'💾 Save Changes\'; }\n  }\n}\n\n/* ── Verify record ── */\nasync function verifyRecord(idx){\n  const t = allTeachers[idx];\n  // Collect any pending edits first\n  const fields = collectEdits(idx);\n  if(Object.keys(fields).length){\n    // Save edits then verify\n    await saveUpdates(idx);\n    // re-read allTeachers[idx] after update\n  }\n\n  const vbtn = document.querySelector(`#card_${idx} .btn-verify`);\n  if(vbtn){ vbtn.disabled=true; vbtn.innerHTML=\'<div class="spinner"></div> Verifying…\'; }\n\n  try{\n    const res  = await fetch(\'/api/teacher/verify\',{\n      method:\'POST\',\n      headers:{\'Content-Type\':\'application/json\'},\n      body: JSON.stringify({teacher_name:t.teacher_name, proforma:t.proforma})\n    });\n    const data = await res.json();\n    if(data.success){\n      t.status = \'VERIFIED\';\n      t.last_updated = data.timestamp || \'\';\n      editModes[cardKey(t)] = false;\n      const card = document.getElementById(`card_${idx}`);\n      card.outerHTML = buildCard(t, idx);\n      showToast(\'✅ \'+data.message, \'success\');\n    } else {\n      showToast(\'❌ \'+(data.error||\'Verification failed.\'), \'error\');\n    }\n  } catch(e){\n    showToast(\'❌ Network error. Please retry.\', \'error\');\n  }\n}\n\n/* ── Toast ── */\nlet toastTimer;\nfunction showToast(msg, type=\'success\'){\n  const el = document.getElementById(\'toast\');\n  el.textContent = msg;\n  el.className = `toast ${type} show`;\n  clearTimeout(toastTimer);\n  toastTimer = setTimeout(()=>{ el.className=\'toast\'; }, 4500);\n}\n\n/* ── Escape HTML ── */\nfunction escHtml(s){\n  if(s==null) return \'\';\n  return String(s).replace(/&/g,\'&amp;\').replace(/</g,\'&lt;\').replace(/>/g,\'&gt;\').replace(/"/g,\'&quot;\');\n}\n</script>\n</body>\n</html>\n'

ADMIN_LOGIN_HTML = '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>Admin Login — Chikiti Block</title>\n<style>\n  :root{--navy:#0d2366;--steel:#1f4e79;--red:#c62828}\n  *{box-sizing:border-box;margin:0;padding:0}\n  body{font-family:\'Segoe UI\',Arial,sans-serif;background:linear-gradient(135deg,#1a1a3e 0%,#0d2366 50%,#1f4e79 100%);min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px}\n  .back-link{position:fixed;top:16px;left:16px;color:rgba(255,255,255,.8);text-decoration:none;font-size:.88rem;font-weight:600;background:rgba(255,255,255,.12);padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,.2)}\n  .back-link:hover{background:rgba(255,255,255,.2)}\n  .card{background:#fff;border-radius:18px;box-shadow:0 16px 60px rgba(0,0,0,.35);padding:44px 40px 40px;max-width:440px;width:100%}\n  .logo{text-align:center;margin-bottom:28px}\n  .logo .icon{font-size:3rem;margin-bottom:10px}\n  .logo h1{font-size:1.45rem;color:var(--navy);font-weight:700}\n  .logo p{font-size:.84rem;color:#666;margin-top:4px}\n  .badge{display:inline-block;background:#0d2366;color:#fff;font-size:.7rem;padding:3px 10px;border-radius:12px;margin-top:6px;letter-spacing:.5px}\n  .field{margin-bottom:18px}\n  label{display:block;font-size:.87rem;font-weight:600;color:var(--steel);margin-bottom:6px}\n  input[type=text],input[type=password]{\n    width:100%;padding:11px 14px;border:2px solid #d0daf0;border-radius:9px;\n    font-size:.95rem;color:#222;outline:none;transition:border-color .2s;background:#fdfdff}\n  input:focus{border-color:var(--navy)}\n  .pw-wrap{position:relative}\n  .pw-toggle{position:absolute;right:12px;top:50%;transform:translateY(-50%);cursor:pointer;font-size:1.1rem;user-select:none}\n  .btn-login{width:100%;padding:13px;background:linear-gradient(135deg,#1a1a3e,var(--navy));color:#fff;border:none;border-radius:10px;font-size:1rem;font-weight:700;cursor:pointer;transition:opacity .2s;margin-top:6px}\n  .btn-login:hover{opacity:.88}\n  .btn-login:disabled{opacity:.6;cursor:not-allowed}\n  .alert{padding:11px 15px;border-radius:8px;font-size:.87rem;margin-bottom:14px;display:none}\n  .alert.error{background:#ffeaea;color:var(--red);border:1px solid #ffcdd2;display:block}\n  .spinner{display:none;width:18px;height:18px;border:3px solid rgba(255,255,255,.4);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;margin:auto}\n  @keyframes spin{to{transform:rotate(360deg)}}\n  .secure-note{text-align:center;margin-top:18px;font-size:.77rem;color:#999}\n</style>\n</head>\n<body>\n\n<a class="back-link" href="/">← Home</a>\n\n<div class="card">\n  <div class="logo">\n    <div class="icon">🛡️</div>\n    <h1>Admin Login</h1>\n    <p>Teacher Verification System</p>\n    <span class="badge">ADMINISTRATOR ACCESS</span>\n  </div>\n\n  <div class="alert error" id="alertBox" style="display:none"></div>\n\n  <div class="field">\n    <label for="username">👤 Username</label>\n    <input type="text" id="username" placeholder="Enter admin username" autocomplete="username">\n  </div>\n\n  <div class="field">\n    <label for="password">🔑 Password</label>\n    <div class="pw-wrap">\n      <input type="password" id="password" placeholder="Enter password" autocomplete="current-password">\n      <span class="pw-toggle" onclick="togglePw()">👁</span>\n    </div>\n  </div>\n\n  <button class="btn-login" id="loginBtn" onclick="doLogin()">\n    <span id="btnText">Login to Admin Panel →</span>\n    <div class="spinner" id="spinner"></div>\n  </button>\n\n  <p class="secure-note">🔒 Access restricted to authorised administrators only</p>\n</div>\n\n<script>\nfunction togglePw(){\n  const pw = document.getElementById(\'password\');\n  pw.type = pw.type===\'password\' ? \'text\' : \'password\';\n}\n\nasync function doLogin(){\n  const username = document.getElementById(\'username\').value.trim();\n  const password = document.getElementById(\'password\').value.trim();\n  const alertBox = document.getElementById(\'alertBox\');\n  alertBox.style.display = \'none\';\n\n  if(!username || !password){ showErr(\'Please enter username and password.\'); return; }\n\n  setBusy(true);\n  try{\n    const res  = await fetch(\'/admin/login\',{\n      method:\'POST\',\n      headers:{\'Content-Type\':\'application/json\'},\n      body: JSON.stringify({username, password})\n    });\n    const data = await res.json();\n    if(data.success){\n      window.location.href = \'/admin/dashboard\';\n    } else {\n      showErr(data.error || \'Invalid credentials.\');\n    }\n  } catch(e){ showErr(\'Network error. Please try again.\'); }\n  finally{ setBusy(false); }\n}\n\nfunction showErr(msg){\n  const a = document.getElementById(\'alertBox\');\n  a.textContent = msg; a.style.display=\'block\';\n}\nfunction setBusy(on){\n  document.getElementById(\'loginBtn\').disabled = on;\n  document.getElementById(\'btnText\').style.display = on?\'none\':\'inline\';\n  document.getElementById(\'spinner\').style.display = on?\'block\':\'none\';\n}\ndocument.addEventListener(\'keydown\', e=>{ if(e.key===\'Enter\') doLogin(); });\n</script>\n</body>\n</html>\n'

ADMIN_DASHBOARD_HTML = '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>Admin Dashboard — Chikiti Block</title>\n<style>\n:root{--navy:#0d2366;--steel:#1f4e79;--sky:#dce6f1;--green:#2e7d32;--amber:#e65100;--red:#c62828;--grey:#f5f7fa;--white:#fff}\n*{box-sizing:border-box;margin:0;padding:0}\nbody{font-family:\'Segoe UI\',Arial,sans-serif;background:var(--grey);min-height:100vh}\n.topbar{background:linear-gradient(135deg,var(--navy),var(--steel));color:#fff;padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:60px}\n.topbar-brand{display:flex;align-items:center;gap:10px;font-size:1.05rem;font-weight:700}\n.topbar-right{display:flex;gap:10px;align-items:center}\n.btn-logout{background:rgba(255,255,255,.12);color:#fff;border:1px solid rgba(255,255,255,.3);padding:7px 16px;border-radius:8px;cursor:pointer;font-size:.85rem;text-decoration:none}\n.btn-logout:hover{background:rgba(255,255,255,.25)}\n.layout{display:flex;min-height:calc(100vh - 60px)}\n.sidebar{width:220px;background:var(--navy);flex-shrink:0;padding:18px 0}\n.sidebar-item{display:flex;align-items:center;gap:10px;padding:12px 20px;color:rgba(255,255,255,.75);cursor:pointer;font-size:.9rem;transition:all .2s}\n.sidebar-item:hover,.sidebar-item.active{background:rgba(255,255,255,.12);color:#fff}\n.sidebar-item .ico{font-size:1.1rem;width:22px;text-align:center}\n.sidebar-sep{border-top:1px solid rgba(255,255,255,.1);margin:8px 0}\n.main{flex:1;padding:24px;overflow-y:auto}\n.tab-pane{display:none}.tab-pane.active{display:block}\n.stats-row{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:16px;margin-bottom:24px}\n.stat-card{background:var(--white);border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.07);padding:18px 20px}\n.stat-card .val{font-size:1.9rem;font-weight:700;color:var(--navy)}\n.stat-card .lbl{font-size:.8rem;color:#777;margin-top:4px}\n.sec-hdr{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;flex-wrap:wrap;gap:10px}\n.sec-hdr h2{font-size:1.1rem;font-weight:700;color:var(--navy)}\n.btn{display:inline-flex;align-items:center;gap:6px;padding:9px 18px;border-radius:8px;font-size:.88rem;font-weight:600;cursor:pointer;border:none;transition:all .2s}\n.btn-primary{background:var(--navy);color:#fff}.btn-primary:hover{opacity:.88}\n.btn-success{background:#2e7d32;color:#fff}.btn-success:hover{opacity:.88}\n.btn-danger{background:var(--red);color:#fff}.btn-danger:hover{opacity:.88}\n.btn-outline{background:#fff;color:var(--navy);border:2px solid var(--navy)}.btn-outline:hover{background:var(--sky)}\n.btn-sm{padding:6px 12px;font-size:.8rem}\n.btn:disabled{opacity:.5;cursor:not-allowed}\n.search-bar{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px;align-items:center}\n.search-bar input,.search-bar select{padding:9px 13px;border:2px solid #dde3f0;border-radius:8px;font-size:.88rem;background:#fff;outline:none}\n.search-bar input:focus,.search-bar select:focus{border-color:var(--navy)}\n.search-bar input{min-width:220px}\n.table-wrap{background:var(--white);border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.07);overflow:hidden}\ntable{width:100%;border-collapse:collapse;font-size:.85rem}\nth{background:var(--navy);color:#fff;padding:11px 13px;text-align:left;font-weight:600;white-space:nowrap}\ntd{padding:10px 13px;border-bottom:1px solid #eef0f6;vertical-align:middle}\ntr:last-child td{border-bottom:none}\ntr:hover td{background:#f8f9ff}\n.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:.75rem;font-weight:700}\n.badge-PENDING{background:#fff3e0;color:#e65100;border:1px solid #ffcc02}\n.badge-VERIFIED{background:#e8f5e9;color:#1b5e20;border:1px solid #66bb6a}\n.badge-UPDATED{background:#e3f2fd;color:#0d47a1;border:1px solid #42a5f5}\n.school-row{display:flex;align-items:center;justify-content:space-between;padding:12px 16px;border-bottom:1px solid #eef0f6;gap:10px}\n.school-row:last-child{border-bottom:none}\n.school-name{font-weight:600;color:var(--navy)}\n.school-meta{font-size:.78rem;color:#888;margin-top:2px}\n.school-actions{display:flex;gap:8px;flex-shrink:0}\n.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;display:none;align-items:center;justify-content:center;padding:20px}\n.modal-overlay.show{display:flex}\n.modal{background:#fff;border-radius:16px;box-shadow:0 12px 48px rgba(0,0,0,.25);width:100%;max-width:700px;max-height:90vh;display:flex;flex-direction:column}\n.modal-lg{max-width:920px}\n.modal-hdr{padding:18px 24px 14px;border-bottom:1px solid #eee;display:flex;align-items:center;justify-content:space-between;flex-shrink:0}\n.modal-hdr h3{font-size:1.05rem;font-weight:700;color:var(--navy)}\n.modal-close{background:none;border:none;font-size:1.4rem;cursor:pointer;color:#999}\n.modal-close:hover{color:#333}\n.modal-body{padding:20px 24px;overflow-y:auto;flex:1}\n.modal-footer{padding:14px 24px 18px;border-top:1px solid #eee;display:flex;gap:10px;justify-content:flex-end;flex-shrink:0}\n.form-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:14px}\n.form-group{display:flex;flex-direction:column;gap:5px}\n.form-group label{font-size:.8rem;font-weight:700;color:var(--steel);text-transform:uppercase;letter-spacing:.3px}\n.form-group input,.form-group select,.form-group textarea{padding:9px 12px;border:2px solid #dde3f0;border-radius:8px;font-size:.9rem;font-family:inherit;outline:none;background:#fdfdff}\n.form-group input:focus,.form-group select:focus{border-color:var(--navy)}\n.form-section-title{grid-column:1/-1;font-size:.92rem;font-weight:700;color:var(--steel);margin:6px 0 2px;padding-bottom:6px;border-bottom:2px solid var(--sky)}\n.toast{position:fixed;bottom:24px;right:24px;padding:13px 20px;border-radius:10px;color:#fff;font-size:.88rem;font-weight:600;z-index:9999;opacity:0;transform:translateY(20px);transition:all .35s;pointer-events:none;max-width:340px;box-shadow:0 4px 20px rgba(0,0,0,.2)}\n.toast.show{opacity:1;transform:translateY(0)}\n.toast.success{background:#2e7d32}.toast.error{background:#c62828}.toast.info{background:#0d2366}\n.pagination{display:flex;gap:6px;justify-content:center;margin-top:14px;flex-wrap:wrap}\n.page-btn{padding:6px 13px;border-radius:7px;border:1.5px solid #dde3f0;background:#fff;cursor:pointer;font-size:.82rem;font-weight:600;color:var(--navy)}\n.page-btn.active,.page-btn:hover{background:var(--navy);color:#fff;border-color:var(--navy)}\n.log-entry{padding:9px 14px;border-bottom:1px solid #eef0f6;font-size:.82rem;color:#444;display:flex;gap:12px}\n.log-entry:last-child{border-bottom:none}\n.log-time{color:#999;white-space:nowrap;font-size:.75rem}\n.log-msg{flex:1}\n.export-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:14px;margin-bottom:18px}\n.export-card{background:#fff;border:2px solid #dde3f0;border-radius:10px;padding:16px;text-align:center;cursor:pointer;transition:all .2s}\n.export-card:hover{border-color:var(--navy);background:#f0f4ff}\n.export-card .ico{font-size:2rem;margin-bottom:8px}\n.export-card .ttl{font-size:.88rem;font-weight:700;color:var(--navy)}\n.export-card .sub{font-size:.76rem;color:#888;margin-top:3px}\n#loadingOverlay{position:fixed;inset:0;background:rgba(255,255,255,.88);display:flex;align-items:center;justify-content:center;z-index:9000;flex-direction:column;gap:14px}\n#loadingOverlay .sp{width:44px;height:44px;border:5px solid #dce6f1;border-top-color:var(--navy);border-radius:50%;animation:spin 1s linear infinite}\n@keyframes spin{to{transform:rotate(360deg)}}\n#loadingOverlay p{color:var(--navy);font-weight:600}\n.confirm-box{background:#fff;border-radius:14px;box-shadow:0 8px 40px rgba(0,0,0,.3);padding:32px;max-width:420px;width:100%;text-align:center}\n.confirm-box h3{font-size:1.1rem;color:var(--navy);margin-bottom:10px}\n.confirm-box p{color:#555;font-size:.9rem;margin-bottom:22px;line-height:1.6}\n.confirm-actions{display:flex;gap:12px;justify-content:center}\n@media(max-width:768px){.sidebar{display:none}.stats-row{grid-template-columns:1fr 1fr}.form-grid{grid-template-columns:1fr}}\n</style>\n</head>\n<body>\n<div id="loadingOverlay"><div class="sp"></div><p>Loading dashboard…</p></div>\n\n<div class="topbar">\n  <div class="topbar-brand">🛡️ Admin Panel — Chikiti Block Teacher Verification</div>\n  <div class="topbar-right">\n    <span style="font-size:.82rem;background:rgba(255,255,255,.15);padding:5px 12px;border-radius:20px">👤 Subrat</span>\n    <a href="/admin/logout" class="btn-logout">🚪 Logout</a>\n  </div>\n</div>\n\n<div class="layout">\n<div class="sidebar">\n  <div class="sidebar-item active" onclick="showTab(\'overview\')" id="nav-overview"><span class="ico">📊</span> Overview</div>\n  <div class="sidebar-item" onclick="showTab(\'teachers\')" id="nav-teachers"><span class="ico">👥</span> Teachers</div>\n  <div class="sidebar-item" onclick="showTab(\'duplicates\')" id="nav-duplicates"><span class="ico">🧬</span> Duplicates</div>\n  <div class="sidebar-item" onclick="showTab(\'schools\')" id="nav-schools"><span class="ico">🏫</span> Schools</div>\n  <div class="sidebar-sep"></div>\n  <div class="sidebar-item" onclick="showTab(\'create\')" id="nav-create"><span class="ico">➕</span> Create Teacher</div>\n  <div class="sidebar-item" onclick="showTab(\'import\')" id="nav-import"><span class="ico">📤</span> Bulk Import</div>\n  <div class="sidebar-item" onclick="showTab(\'export\')" id="nav-export"><span class="ico">📥</span> Export</div>\n  <div class="sidebar-item" onclick="showTab(\'logs\')" id="nav-logs"><span class="ico">📋</span> Activity Log</div>\n</div>\n\n<div class="main">\n\n<!-- OVERVIEW -->\n<div class="tab-pane active" id="tab-overview">\n  <h2 style="color:var(--navy);margin-bottom:18px">📊 Dashboard Overview</h2>\n  <div class="stats-row">\n    <div class="stat-card"><div class="val" id="statTotal">—</div><div class="lbl">Total Teachers</div></div>\n    <div class="stat-card"><div class="val" id="statVerified" style="color:#2e7d32">—</div><div class="lbl">Verified</div></div>\n    <div class="stat-card"><div class="val" id="statUpdated" style="color:#1565c0">—</div><div class="lbl">Updated</div></div>\n    <div class="stat-card"><div class="val" id="statPending" style="color:#e65100">—</div><div class="lbl">Pending</div></div>\n    <div class="stat-card"><div class="val" id="statSchools" style="color:#6a1b9a">—</div><div class="lbl">Schools</div></div>\n    <div class="stat-card"><div class="val" id="statDuplicates" style="color:#ad1457">—</div><div class="lbl">Duplicate Groups</div></div>\n  </div>\n  <div class="table-wrap">\n    <div style="padding:14px 18px;border-bottom:1px solid #eee;font-weight:700;color:var(--navy)">📋 Proforma Summary</div>\n    <table><thead><tr><th>Proforma</th><th>Total</th><th>Verified</th><th>Updated</th><th>Pending</th><th>Progress</th></tr></thead>\n    <tbody id="proformaSummary"></tbody></table>\n  </div>\n</div>\n\n<!-- TEACHERS -->\n<div class="tab-pane" id="tab-teachers">\n  <div class="sec-hdr">\n    <h2>👥 Teacher Records</h2>\n    <button class="btn btn-success" onclick="showTab(\'create\')">➕ Add New Teacher</button>\n  </div>\n  <div class="search-bar">\n    <input type="text" id="searchInput" placeholder="🔍 Search name / school…" oninput="filterTeachers()">\n    <select id="filterProforma" onchange="filterTeachers()">\n      <option value="">All Proformas</option>\n      <option>PROFORMA I</option><option>PROFORMA II</option><option>PROFORMA III</option>\n      <option>PROFORMA IV</option><option>PROFORMA V</option><option>PROFORMA VI</option><option>PROFORMA VII</option><option>PROFORMA VIII</option>\n    </select>\n    <select id="filterStatus" onchange="filterTeachers()">\n      <option value="">All Status</option><option>PENDING</option><option>VERIFIED</option><option>UPDATED</option>\n    </select>\n    <select id="filterSchool" onchange="filterTeachers()"><option value="">All Schools</option></select>\n  </div>\n  <div class="table-wrap">\n    <table>\n      <thead><tr><th>#</th><th>Name</th><th>School</th><th>Proforma</th><th>Status</th><th>Last Updated</th><th style="text-align:center">Actions</th></tr></thead>\n      <tbody id="teacherTableBody"></tbody>\n    </table>\n  </div>\n  <div class="pagination" id="teacherPagination"></div>\n  <div style="font-size:.8rem;color:#999;margin-top:8px;text-align:right" id="teacherCount"></div>\n</div>\n\n<!-- DUPLICATES -->\n<div class="tab-pane" id="tab-duplicates">\n  <div class="sec-hdr">\n    <h2>🧬 Duplicate Teachers Across Proformas</h2>\n    <button class="btn btn-outline btn-sm" onclick="loadDuplicates()">🔄 Refresh</button>\n  </div>\n  <div style="font-size:.84rem;color:#666;margin-bottom:12px">This section shows teachers appearing in more than one proforma. You can directly edit or delete any listed record.</div>\n  <div class="table-wrap">\n    <table>\n      <thead><tr><th>Group</th><th>Teacher</th><th>School</th><th>DOB</th><th>Proforma</th><th>Status</th><th style="text-align:center">Actions</th></tr></thead>\n      <tbody id="duplicateTableBody"></tbody>\n    </table>\n  </div>\n  <div style="font-size:.8rem;color:#999;margin-top:8px;text-align:right" id="duplicateCount"></div>\n</div>\n\n<!-- SCHOOLS -->\n<div class="tab-pane" id="tab-schools">\n  <div class="sec-hdr">\n    <h2>🏫 School Management</h2>\n    <button class="btn btn-success" onclick="openAddSchool()">➕ Add New School</button>\n  </div>\n  <div class="table-wrap" id="schoolList">\n    <div style="padding:20px;text-align:center;color:#999">Loading schools…</div>\n  </div>\n</div>\n\n<!-- CREATE TEACHER -->\n<div class="tab-pane" id="tab-create">\n  <div class="sec-hdr"><h2>➕ Create New Teacher</h2></div>\n  <div style="background:#fff;border-radius:14px;box-shadow:0 2px 12px rgba(0,0,0,.07);padding:24px">\n    <div class="form-grid">\n      <div class="form-section-title">🏫 Basic Information</div>\n      <div class="form-group"><label>Proforma *</label>\n        <select id="ct_proforma"><option value="">— Select —</option>\n        <option>PROFORMA I</option><option>PROFORMA II</option><option>PROFORMA III</option>\n        <option>PROFORMA IV</option><option>PROFORMA V</option><option>PROFORMA VI</option><option>PROFORMA VII</option><option>PROFORMA VIII</option></select>\n      </div>\n      <div class="form-group"><label>Teacher Name *</label><input type="text" id="ct_teacher_name" placeholder="Full name"></div>\n      <div class="form-group"><label>School *</label>\n        <select id="ct_school"><option value="">— Select School —</option></select>\n      </div>\n      <div class="form-group"><label>District / Block</label><input type="text" id="ct_district" value="Ganjam/Chikiti"></div>\n      <div class="form-group"><label>S.No. (auto if blank)</label><input type="text" id="ct_sl_no" placeholder="e.g. 1"></div>\n      <div class="form-section-title">🎓 Qualifications</div>\n      <div class="form-group"><label>General Qualification</label><input type="text" id="ct_general_qualification" placeholder="e.g. B.A."></div>\n      <div class="form-group"><label>Date of Result (General)</label><input type="text" id="ct_date_general_result" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Passing (General)</label><input type="text" id="ct_date_passing" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Training Qualification</label><input type="text" id="ct_training_qualification" placeholder="e.g. B.Ed."></div>\n      <div class="form-group"><label>Date of Result (Training)</label><input type="text" id="ct_date_training_result" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Passing (Training)</label><input type="text" id="ct_date_training_passing" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>OTET (I/II/NO)</label>\n        <select id="ct_otet"><option value="">—</option><option>I</option><option>II</option><option>I &amp; II</option><option>NO</option></select>\n      </div>\n      <div class="form-group"><label>Date of Passing OTET</label><input type="text" id="ct_date_otet" placeholder="DD/MM/YYYY"></div>\n      <div class="form-section-title">📋 Personal Details</div>\n      <div class="form-group"><label>Category</label>\n        <select id="ct_category"><option value="">—</option><option>GEN</option><option>OBC</option><option>SC</option><option>ST</option><option>EWS</option></select>\n      </div>\n      <div class="form-group"><label>Date of Birth *</label><input type="text" id="ct_dob" placeholder="DD/MM/YYYY"></div>\n      <div class="form-section-title">📅 Service Details</div>\n      <div class="form-group"><label>Date of 1st Appointment (SS)</label><input type="text" id="ct_date_first_appt_ss" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Regular Appointment</label><input type="text" id="ct_date_regular_teacher" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Regularisation for Consideration</label><input type="text" id="ct_date_regularisation_consideration" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Notification No. &amp; Date</label><input type="text" id="ct_notification_no"></div>\n      <div class="form-group"><label>Rank in Merit List</label><input type="text" id="ct_rank"></div>\n      <div class="form-group"><label>Date of Joining Level-V(B)</label><input type="text" id="ct_date_joining_levelVB" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Joining Level-III</label><input type="text" id="ct_date_joining_levelIII" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Joining Level-IV</label><input type="text" id="ct_date_joining_levelIV" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>Date of Superannuation</label><input type="text" id="ct_date_superannuation" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>IDT / RA</label><input type="text" id="ct_idt_ra"></div>\n      <div class="form-group"><label>Date of Joining District</label><input type="text" id="ct_date_joining_district" placeholder="DD/MM/YYYY"></div>\n      <div class="form-group"><label>DP/CP/Vigilance Pending?</label>\n        <select id="ct_dp_cp_vigilance"><option value="">—</option><option>NO</option><option>YES</option><option>DP</option><option>CP</option><option>Vigilance</option></select>\n      </div>\n      <div class="form-group"><label>Option</label><input type="text" id="ct_option"></div>\n      <div class="form-group"><label>Remark</label><input type="text" id="ct_remark"></div>\n    </div>\n    <div style="margin-top:22px;display:flex;gap:12px;justify-content:flex-end">\n      <button class="btn btn-outline" onclick="clearCreateForm()">🗑️ Clear</button>\n      <button class="btn btn-success" onclick="submitCreateTeacher()" id="createBtn">✅ Create Teacher Record</button>\n    </div>\n  </div>\n</div>\n\n<!-- BULK IMPORT -->\n<div class="tab-pane" id="tab-import">\n  <div class="sec-hdr"><h2>📤 Bulk Import Teachers from Excel</h2></div>\n  <div style="background:#fff;border-radius:14px;box-shadow:0 2px 12px rgba(0,0,0,.07);padding:24px">\n    <div style="font-size:.88rem;color:#555;line-height:1.7;margin-bottom:16px">\n      Upload an Excel file and import multiple teachers directly into a selected proforma. Existing teachers with the same <strong>name + proforma</strong> will be updated automatically; new ones will be added.\n    </div>\n    <div class="form-grid">\n      <div class="form-group">\n        <label>Excel File *</label>\n        <input type="file" id="bulkFile" accept=".xlsx,.xlsm">\n      </div>\n      <div class="form-group">\n        <label>Target Proforma *</label>\n        <select id="imp_proforma">\n          <option value="">— Select —</option>\n          <option>PROFORMA I</option><option>PROFORMA II</option><option>PROFORMA III</option>\n          <option>PROFORMA IV</option><option>PROFORMA V</option><option>PROFORMA VI</option>\n          <option>PROFORMA VII</option><option>PROFORMA VIII</option>\n        </select>\n      </div>\n      <div class="form-group">\n        <label>Default School (optional)</label>\n        <select id="imp_default_school"><option value="">— Use school from Excel / None —</option></select>\n      </div>\n    </div>\n    <div style="margin-top:14px;padding:14px 16px;background:#f7faff;border-left:4px solid #0d2366;border-radius:8px;font-size:.84rem;color:#334">\n      <strong>Supported headers</strong>: Teacher Name, School Name, District/Block, DOB, Category, General Qualification, Training Qualification, OTET, Remark, and Proforma-specific dates.\nFor Proforma VIII, the importer also understands headers like <em>NAME OF THE UPS HM</em>, <em>DATE OF JOINING IN LEVEL-III POST</em>, and <em>DATE OF REGULARISATION FOR CONSIDERATION</em>.\n    </div>\n    <div style="margin-top:18px;display:flex;gap:12px;align-items:center;flex-wrap:wrap">\n      <button class="btn btn-primary" onclick="submitBulkImport()" id="importBtn">📤 Import Excel</button>\n      <span style="font-size:.82rem;color:#777">Accepted format: .xlsx / .xlsm</span>\n    </div>\n    <div id="importResult" style="display:none;margin-top:18px;padding:16px;border-radius:10px;background:#f8f9ff;border:1px solid #dde3f0"></div>\n  </div>\n</div>\n\n<!-- EXPORT -->\n<div class="tab-pane" id="tab-export">\n  <div class="sec-hdr"><h2>📥 Export Data</h2></div>\n  <div class="export-grid">\n    <div class="export-card" onclick="doExport(\'all\',\'all\')"><div class="ico">📊</div><div class="ttl">All Teachers</div><div class="sub">All proformas, all status</div></div>\n    <div class="export-card" onclick="doExport(\'all\',\'PENDING\')"><div class="ico">⏳</div><div class="ttl">Pending Only</div><div class="sub">Not yet verified</div></div>\n    <div class="export-card" onclick="doExport(\'all\',\'VERIFIED\')"><div class="ico">✅</div><div class="ttl">Verified Only</div><div class="sub">Verified records</div></div>\n    <div class="export-card" onclick="doExport(\'all\',\'UPDATED\')"><div class="ico">🔄</div><div class="ttl">Updated Only</div><div class="sub">Updated by teachers</div></div>\n  </div>\n  <div style="background:#fff;border-radius:12px;padding:18px 20px;box-shadow:0 2px 10px rgba(0,0,0,.07)">\n    <div style="font-weight:700;color:var(--navy);margin-bottom:12px">Custom Export</div>\n    <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end">\n      <div><label style="font-size:.8rem;font-weight:600;display:block;margin-bottom:5px">Proforma</label>\n        <select id="expProforma" style="padding:9px 12px;border:2px solid #dde3f0;border-radius:8px;font-size:.88rem">\n          <option value="all">All Proformas</option>\n          <option>PROFORMA I</option><option>PROFORMA II</option><option>PROFORMA III</option>\n          <option>PROFORMA IV</option><option>PROFORMA V</option><option>PROFORMA VI</option><option>PROFORMA VII</option><option>PROFORMA VIII</option>\n        </select></div>\n      <div><label style="font-size:.8rem;font-weight:600;display:block;margin-bottom:5px">Status</label>\n        <select id="expStatus" style="padding:9px 12px;border:2px solid #dde3f0;border-radius:8px;font-size:.88rem">\n          <option value="all">All Status</option><option>PENDING</option><option>VERIFIED</option><option>UPDATED</option>\n        </select></div>\n      <button class="btn btn-primary" onclick="doExport(document.getElementById(\'expProforma\').value,document.getElementById(\'expStatus\').value)">📥 Export Excel</button>\n    </div>\n  </div>\n</div>\n\n<!-- LOGS -->\n<div class="tab-pane" id="tab-logs">\n  <div class="sec-hdr"><h2>📋 Activity Log</h2><button class="btn btn-outline btn-sm" onclick="loadLogs()">🔄 Refresh</button></div>\n  <div class="table-wrap" id="logContainer"><div style="padding:20px;text-align:center;color:#999">Loading…</div></div>\n</div>\n\n</div><!-- /main -->\n</div><!-- /layout -->\n\n<div class="toast" id="toast"></div>\n\n<!-- Edit Modal -->\n<div class="modal-overlay" id="editModal">\n  <div class="modal modal-lg">\n    <div class="modal-hdr"><h3 id="editModalTitle">✏️ Edit Teacher</h3><button class="modal-close" onclick="closeModal(\'editModal\')">✕</button></div>\n    <div class="modal-body"><div class="form-grid" id="editForm"></div></div>\n    <div class="modal-footer">\n      <button class="btn btn-outline" onclick="closeModal(\'editModal\')">Cancel</button>\n      <button class="btn btn-primary" onclick="saveEditTeacher()" id="saveEditBtn">💾 Save Changes</button>\n    </div>\n  </div>\n</div>\n\n<!-- Add/Edit School Modal -->\n<div class="modal-overlay" id="addSchoolModal">\n  <div class="modal" style="max-width:450px">\n    <div class="modal-hdr"><h3 id="addSchoolTitle">🏫 Add School</h3><button class="modal-close" onclick="closeModal(\'addSchoolModal\')">✕</button></div>\n    <div class="modal-body">\n      <div class="form-group">\n        <label>School Name *</label>\n        <input type="text" id="schoolNameInput" placeholder="Enter full school name" style="width:100%">\n      </div>\n      <input type="hidden" id="schoolOldName">\n    </div>\n    <div class="modal-footer">\n      <button class="btn btn-outline" onclick="closeModal(\'addSchoolModal\')">Cancel</button>\n      <button class="btn btn-success" onclick="submitSchool()" id="submitSchoolBtn">✅ Save</button>\n    </div>\n  </div>\n</div>\n\n<!-- Confirm Modal -->\n<div class="modal-overlay" id="confirmModal">\n  <div class="confirm-box">\n    <h3 id="confirmTitle">Confirm</h3>\n    <p id="confirmMsg">Are you sure?</p>\n    <div class="confirm-actions">\n      <button class="btn btn-outline" onclick="closeModal(\'confirmModal\')">Cancel</button>\n      <button class="btn btn-danger" id="confirmOkBtn">Confirm</button>\n    </div>\n  </div>\n</div>\n\n<script>\nconst FIELD_LABELS = {"sl_no": "S.No.", "district": "District / Block", "teacher_name": "Name of Teacher", "school": "School Name", "general_qualification": "General Qualification", "date_general_result": "Date of Result (General)", "date_passing": "Date of Passing (General)", "training_qualification": "Training Qualification", "date_training_result": "Date of Result (Training)", "date_training_passing": "Date of Passing (Training)", "otet": "OTET (I / II / NO)", "date_otet": "Date of Passing OTET", "category": "Category", "dob": "Date of Birth", "date_first_appt_ss": "Date of 1st Appointment (SS)", "date_regular_teacher": "Date of Regular Appointment", "notification_no": "Notification No. & Date", "rank": "Rank in Merit List", "date_joining_levelVB": "Date of Joining Level-V(B)", "date_joining_levelIV": "Date of Joining Level-IV", "date_superannuation": "Date of Superannuation", "idt_ra": "IDT / RA", "date_joining_district": "Date of Joining District", "dp_cp_vigilance": "DP/CP/Vigilance Pending?", "option": "Option", "remark": "Remark"};\nconst FIELD_GROUPS = [{"title": "\\ud83c\\udfeb School & Location", "fields": ["district", "school"]}, {"title": "\\ud83c\\udf93 Qualifications", "fields": ["general_qualification", "date_general_result", "date_passing", "training_qualification", "date_training_result", "date_training_passing", "otet", "date_otet"]}, {"title": "\\ud83d\\udccb Personal Details", "fields": ["category", "dob"]}, {"title": "\\ud83d\\udcc5 Service Details", "fields": ["date_first_appt_ss", "date_regular_teacher", "notification_no", "rank", "date_joining_levelVB", "date_joining_levelIV", "date_superannuation", "idt_ra", "date_joining_district", "dp_cp_vigilance", "option", "remark"]}];\n\nlet allTeachers=[], allSchools=[], filteredTeachers=[], duplicateGroups=[];\nlet currentPage=1; const PAGE_SIZE=20;\nlet editingTeacher=null;\n\nwindow.onload = async()=>{\n  await Promise.all([loadStats(), loadTeachers(), loadSchools(), loadDuplicates()]);\n  document.getElementById(\'loadingOverlay\').style.display=\'none\';\n  populateSchoolDropdowns();\n};\n\nfunction showTab(name){\n  document.querySelectorAll(\'.tab-pane\').forEach(p=>p.classList.remove(\'active\'));\n  document.querySelectorAll(\'.sidebar-item\').forEach(i=>i.classList.remove(\'active\'));\n  document.getElementById(\'tab-\'+name).classList.add(\'active\');\n  document.getElementById(\'nav-\'+name).classList.add(\'active\');\n  if(name===\'logs\') loadLogs();\n  if(name===\'schools\') renderSchools();\n  if(name===\'duplicates\') loadDuplicates();\n  if(name===\'create\' || name===\'import\') populateSchoolDropdowns();\n}\n\nasync function loadStats(){\n  try{\n    const r=await fetch(\'/api/admin/stats\');\n    const d=await r.json();\n    document.getElementById(\'statTotal\').textContent=d.total||0;\n    document.getElementById(\'statVerified\').textContent=d.verified||0;\n    document.getElementById(\'statUpdated\').textContent=d.updated||0;\n    document.getElementById(\'statPending\').textContent=d.pending||0;\n    document.getElementById(\'statSchools\').textContent=d.schools||0;\n    const dupEl = document.getElementById(\'statDuplicates\'); if(dupEl) dupEl.textContent = d.duplicate_groups||0;\n    const tbody=document.getElementById(\'proformaSummary\');\n    tbody.innerHTML=\'\';\n    for(const [pf,s] of Object.entries(d.by_proforma||{})){\n      const pct=s.total?Math.round((s.verified+s.updated)/s.total*100):0;\n      tbody.innerHTML+=`<tr><td><strong>${pf}</strong></td><td>${s.total}</td>\n        <td style="color:#2e7d32">${s.verified}</td><td style="color:#1565c0">${s.updated}</td>\n        <td style="color:#e65100">${s.pending}</td>\n        <td><div style="background:#eee;border-radius:4px;height:8px;width:120px;display:inline-block"><div style="background:#0d2366;height:8px;border-radius:4px;width:${pct}%"></div></div> <span style="font-size:.75rem;color:#666">${pct}%</span></td>\n      </tr>`;\n    }\n  }catch(e){console.error(e);}\n}\n\nasync function loadTeachers(){\n  try{\n    const r=await fetch(\'/api/admin/teachers?limit=9999\');\n    const d=await r.json();\n    allTeachers=d.teachers||[];\n    const schools=[...new Set(allTeachers.map(t=>t.school).filter(Boolean))].sort();\n    const sel=document.getElementById(\'filterSchool\');\n    sel.innerHTML=\'<option value="">All Schools</option>\';\n    schools.forEach(s=>{sel.innerHTML+=`<option>${esc(s)}</option>`;});\n    filteredTeachers=[...allTeachers];\n    renderTeacherTable();\n  }catch(e){console.error(e);}\n}\n\nfunction filterTeachers(){\n  const q=document.getElementById(\'searchInput\').value.toLowerCase();\n  const pf=document.getElementById(\'filterProforma\').value;\n  const st=document.getElementById(\'filterStatus\').value;\n  const sc=document.getElementById(\'filterSchool\').value;\n  filteredTeachers=allTeachers.filter(t=>{\n    if(pf&&t.proforma!==pf) return false;\n    if(st&&t.status!==st) return false;\n    if(sc&&t.school!==sc) return false;\n    if(q&&!`${t.teacher_name} ${t.school}`.toLowerCase().includes(q)) return false;\n    return true;\n  });\n  currentPage=1;\n  renderTeacherTable();\n}\n\nfunction renderTeacherTable(){\n  const total=filteredTeachers.length;\n  const pages=Math.ceil(total/PAGE_SIZE)||1;\n  if(currentPage>pages) currentPage=pages;\n  const slice=filteredTeachers.slice((currentPage-1)*PAGE_SIZE,currentPage*PAGE_SIZE);\n  const tbody=document.getElementById(\'teacherTableBody\');\n  tbody.innerHTML=\'\';\n  slice.forEach((t,i)=>{\n    const gi=(currentPage-1)*PAGE_SIZE+i+1;\n    tbody.innerHTML+=`<tr>\n      <td style="color:#999;font-size:.8rem">${gi}</td>\n      <td><strong>${esc(t.teacher_name)}</strong></td>\n      <td style="font-size:.82rem">${esc(t.school)}</td>\n      <td style="font-size:.8rem;color:#555">${t.proforma}</td>\n      <td><span class="badge badge-${t.status||\'PENDING\'}">${t.status||\'PENDING\'}</span></td>\n      <td style="font-size:.78rem;color:#888">${t.last_updated||\'—\'}</td>\n      <td style="text-align:center;white-space:nowrap">\n        <button class="btn btn-outline btn-sm" onclick="openEditTeacher(\'${esc(t.teacher_name)}\',\'${esc(t.proforma)}\')">✏️ Edit</button>\n        <button class="btn btn-danger btn-sm" style="margin-left:4px" onclick="confirmDeleteTeacher(\'${esc(t.teacher_name)}\',\'${esc(t.proforma)}\')">🗑️</button>\n      </td>\n    </tr>`;\n  });\n  if(!slice.length) tbody.innerHTML=\'<tr><td colspan="7" style="text-align:center;color:#999;padding:24px">No records found.</td></tr>\';\n  document.getElementById(\'teacherCount\').textContent=`Showing ${slice.length} of ${total} records`;\n  renderPagination(pages);\n}\n\nfunction renderPagination(pages){\n  const pg=document.getElementById(\'teacherPagination\');\n  if(pages<=1){pg.innerHTML=\'\';return;}\n  let html=\'\';\n  for(let p=1;p<=pages;p++) html+=`<button class="page-btn ${p===currentPage?\'active\':\'\'}" onclick="goPage(${p})">${p}</button>`;\n  pg.innerHTML=html;\n}\nfunction goPage(p){currentPage=p;renderTeacherTable();}\n\n// Edit Teacher\nfunction openEditTeacher(name,proforma){\n  const t=allTeachers.find(x=>x.teacher_name===name&&x.proforma===proforma);\n  if(!t){showToast(\'Teacher not found\',\'error\');return;}\n  editingTeacher=t;\n  document.getElementById(\'editModalTitle\').textContent=`✏️ Edit: ${name}`;\n  const PROTECTED=[\'sl_no\',\'proforma\',\'change_history\'];\n  let html=`<div class="form-section-title">🔒 System Fields</div>\n    <div class="form-group"><label>S.No.</label><input type="text" id="ef_sl_no" value="${esc(t.sl_no||\'\')}" disabled style="background:#f5f5f5;color:#999"></div>\n    <div class="form-group"><label>Teacher Name</label><input type="text" id="ef_teacher_name" value="${esc(t.teacher_name||\'\')}" disabled style="background:#f5f5f5;color:#999"></div>\n    <div class="form-group"><label>Proforma</label><input type="text" id="ef_proforma_display" value="${esc(t.proforma||\'\')}" disabled style="background:#f5f5f5;color:#999"></div>\n    <div class="form-group"><label>Status</label>\n      <select id="ef_status">\n        <option ${t.status===\'PENDING\'?\'selected\':\'\'}>PENDING</option>\n        <option ${t.status===\'VERIFIED\'?\'selected\':\'\'}>VERIFIED</option>\n        <option ${t.status===\'UPDATED\'?\'selected\':\'\'}>UPDATED</option>\n      </select></div>`;\n  for(const grp of FIELD_GROUPS){\n    html+=`<div class="form-section-title">${grp.title}</div>`;\n    for(const fk of grp.fields){\n      if(fk===\'teacher_name\') continue;\n      const label=FIELD_LABELS[fk]||fk;\n      const val=t[fk]||\'\';\n      if(fk===\'school\'){\n        let opts=allSchools.map(s=>`<option ${s===val?\'selected\':\'\'}>${esc(s)}</option>`).join(\'\');\n        html+=`<div class="form-group"><label>${label}</label><select id="ef_${fk}">${opts}</select></div>`;\n      } else {\n        html+=`<div class="form-group"><label>${label}</label><input type="text" id="ef_${fk}" value="${esc(val)}" placeholder="${fk.startsWith(\'date_\')||fk===\'dob\'?\'DD/MM/YYYY\':\'\'}"></div>`;\n      }\n    }\n  }\n  document.getElementById(\'editForm\').innerHTML=html;\n  openModal(\'editModal\');\n}\n\nasync function saveEditTeacher(){\n  if(!editingTeacher) return;\n  const updates={};\n  for(const fk of Object.keys(FIELD_LABELS)){\n    if([\'sl_no\',\'teacher_name\',\'proforma\'].includes(fk)) continue;\n    const el=document.getElementById(\'ef_\'+fk);\n    if(el&&!el.disabled) updates[fk]=el.value.trim();\n  }\n  const stEl=document.getElementById(\'ef_status\');\n  if(stEl) updates[\'status\']=stEl.value;\n  const btn=document.getElementById(\'saveEditBtn\');\n  btn.disabled=true;btn.textContent=\'⏳ Saving…\';\n  try{\n    const r=await fetch(\'/api/admin/teacher/update\',{\n      method:\'POST\',headers:{\'Content-Type\':\'application/json\'},\n      body:JSON.stringify({teacher_name:editingTeacher.teacher_name,proforma:editingTeacher.proforma,updates})\n    });\n    const d=await r.json();\n    if(d.success){showToast(\'Teacher updated!\',\'success\');closeModal(\'editModal\');await loadTeachers();await loadStats();await loadDuplicates();}\n    else showToast(d.error||\'Update failed\',\'error\');\n  }catch(e){showToast(\'Network error\',\'error\');}\n  finally{btn.disabled=false;btn.textContent=\'💾 Save Changes\';}\n}\n\nfunction confirmDeleteTeacher(name,proforma){\n  document.getElementById(\'confirmTitle\').textContent=\'🗑️ Delete Teacher\';\n  document.getElementById(\'confirmMsg\').textContent=`Delete "${name}" (${proforma})? This cannot be undone.`;\n  document.getElementById(\'confirmOkBtn\').onclick=()=>deleteTeacher(name,proforma);\n  openModal(\'confirmModal\');\n}\nasync function deleteTeacher(name,proforma){\n  closeModal(\'confirmModal\');\n  try{\n    const r=await fetch(\'/api/admin/teacher/delete\',{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify({teacher_name:name,proforma})});\n    const d=await r.json();\n    if(d.success){showToast(\'Teacher deleted.\',\'success\');await loadTeachers();await loadStats();await loadDuplicates();}\n    else showToast(d.error||\'Failed\',\'error\');\n  }catch(e){showToast(\'Network error\',\'error\');}\n}\n\n// Schools\nasync function loadSchools(){\n  try{\n    const r=await fetch(\'/api/admin/schools\');\n    const d=await r.json();\n    allSchools=d.schools||[];\n    document.getElementById(\'statSchools\').textContent=allSchools.length;\n  }catch(e){console.error(e);}\n}\n\nfunction renderSchools(){\n  const wrap=document.getElementById(\'schoolList\');\n  if(!allSchools.length){wrap.innerHTML=\'<div style="padding:20px;text-align:center;color:#999">No schools found.</div>\';return;}\n  let html=\'\';\n  allSchools.forEach(s=>{\n    const cnt=allTeachers.filter(t=>t.school===s).length;\n    html+=`<div class="school-row">\n      <div><div class="school-name">${esc(s)}</div><div class="school-meta">${cnt} teacher${cnt!==1?\'s\':\'\'}</div></div>\n      <div class="school-actions">\n        <button class="btn btn-outline btn-sm" onclick="openEditSchool(\'${esc(s)}\')">✏️ Rename</button>\n        <button class="btn btn-danger btn-sm" onclick="confirmDeleteSchool(\'${esc(s)}\',${cnt})">🗑️ Delete</button>\n      </div>\n    </div>`;\n  });\n  wrap.innerHTML=html;\n}\n\nfunction openAddSchool(){\n  document.getElementById(\'addSchoolTitle\').textContent=\'🏫 Add New School\';\n  document.getElementById(\'schoolNameInput\').value=\'\';\n  document.getElementById(\'schoolOldName\').value=\'\';\n  document.getElementById(\'submitSchoolBtn\').textContent=\'✅ Add School\';\n  openModal(\'addSchoolModal\');\n  setTimeout(()=>document.getElementById(\'schoolNameInput\').focus(),100);\n}\nfunction openEditSchool(name){\n  document.getElementById(\'addSchoolTitle\').textContent=\'✏️ Rename School\';\n  document.getElementById(\'schoolNameInput\').value=name;\n  document.getElementById(\'schoolOldName\').value=name;\n  document.getElementById(\'submitSchoolBtn\').textContent=\'✅ Save Name\';\n  openModal(\'addSchoolModal\');\n  setTimeout(()=>document.getElementById(\'schoolNameInput\').focus(),100);\n}\n\nasync function submitSchool(){\n  const name=document.getElementById(\'schoolNameInput\').value.trim();\n  const oldName=document.getElementById(\'schoolOldName\').value.trim();\n  if(!name){showToast(\'Please enter a school name\',\'error\');return;}\n  const btn=document.getElementById(\'submitSchoolBtn\');\n  btn.disabled=true;\n  try{\n    let endpoint,body;\n    if(oldName&&oldName!==name){endpoint=\'/api/admin/school/rename\';body={old_name:oldName,new_name:name};}\n    else if(!oldName){endpoint=\'/api/admin/school/add\';body={name};}\n    else{closeModal(\'addSchoolModal\');return;}\n    const r=await fetch(endpoint,{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify(body)});\n    const d=await r.json();\n    if(d.success){\n      showToast(d.message||\'Done!\',\'success\');\n      closeModal(\'addSchoolModal\');\n      await loadSchools();\n      await loadTeachers();\n      renderSchools();\n      populateSchoolDropdowns();\n    }else showToast(d.error||\'Failed\',\'error\');\n  }catch(e){showToast(\'Network error\',\'error\');}\n  finally{btn.disabled=false;}\n}\n\nfunction confirmDeleteSchool(name,count){\n  const msg=count>0?`"${name}" has ${count} teacher(s). Remove this school? Teachers won\'t be deleted.`:`Delete school "${name}"?`;\n  document.getElementById(\'confirmTitle\').textContent=\'🗑️ Delete School\';\n  document.getElementById(\'confirmMsg\').textContent=msg;\n  document.getElementById(\'confirmOkBtn\').onclick=()=>deleteSchool(name);\n  openModal(\'confirmModal\');\n}\nasync function deleteSchool(name){\n  closeModal(\'confirmModal\');\n  try{\n    const r=await fetch(\'/api/admin/school/delete\',{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify({name})});\n    const d=await r.json();\n    if(d.success){showToast(\'School removed.\',\'success\');await loadSchools();renderSchools();populateSchoolDropdowns();}\n    else showToast(d.error||\'Failed\',\'error\');\n  }catch(e){showToast(\'Network error\',\'error\');}\n}\n\nfunction populateSchoolDropdowns(){\n  const configs = [\n    {id:\'ct_school\', empty:\'— Select School —\'},\n    {id:\'imp_default_school\', empty:\'— Use school from Excel / None —\'}\n  ];\n  configs.forEach(cfg=>{\n    const sel=document.getElementById(cfg.id);\n    if(!sel) return;\n    const cur=sel.value;\n    sel.innerHTML=`<option value="">${cfg.empty}</option>`;\n    allSchools.forEach(s=>{sel.innerHTML+=`<option ${s===cur?\'selected\':\'\'}>${esc(s)}</option>`;});\n  });\n}\n\n// Create Teacher\nconst CT_FIELDS=[\'ct_proforma\',\'ct_teacher_name\',\'ct_school\',\'ct_district\',\'ct_sl_no\',\n  \'ct_general_qualification\',\'ct_date_general_result\',\'ct_date_passing\',\n  \'ct_training_qualification\',\'ct_date_training_result\',\'ct_date_training_passing\',\n  \'ct_otet\',\'ct_date_otet\',\'ct_category\',\'ct_dob\',\'ct_date_first_appt_ss\',\n  \'ct_date_regular_teacher\',\'ct_date_regularisation_consideration\',\'ct_notification_no\',\'ct_rank\',\'ct_date_joining_levelVB\',\n  \'ct_date_joining_levelIII\',\'ct_date_joining_levelIV\',\'ct_date_superannuation\',\'ct_idt_ra\',\'ct_date_joining_district\',\n  \'ct_dp_cp_vigilance\',\'ct_option\',\'ct_remark\'];\n\nfunction clearCreateForm(){\n  CT_FIELDS.forEach(id=>{\n    const el=document.getElementById(id);\n    if(!el) return;\n    if(el.tagName===\'SELECT\') el.value=\'\';\n    else el.value=id===\'ct_district\'?\'Ganjam/Chikiti\':\'\';\n  });\n}\n\nasync function submitCreateTeacher(){\n  const proforma=document.getElementById(\'ct_proforma\').value;\n  const teacher_name=document.getElementById(\'ct_teacher_name\').value.trim();\n  const school=document.getElementById(\'ct_school\').value;\n  const dob=document.getElementById(\'ct_dob\').value.trim();\n  if(!proforma){showToast(\'Please select a proforma\',\'error\');return;}\n  if(!teacher_name){showToast(\'Teacher name is required\',\'error\');return;}\n  if(!school){showToast(\'Please select a school\',\'error\');return;}\n  if(!dob){showToast(\'Date of birth is required\',\'error\');return;}\n\n  const record={proforma,teacher_name,school,dob,\n    district:document.getElementById(\'ct_district\').value.trim()||\'Ganjam/Chikiti\',\n    sl_no:document.getElementById(\'ct_sl_no\').value.trim(),\n    general_qualification:document.getElementById(\'ct_general_qualification\').value.trim(),\n    date_general_result:document.getElementById(\'ct_date_general_result\').value.trim(),\n    date_passing:document.getElementById(\'ct_date_passing\').value.trim(),\n    training_qualification:document.getElementById(\'ct_training_qualification\').value.trim(),\n    date_training_result:document.getElementById(\'ct_date_training_result\').value.trim(),\n    date_training_passing:document.getElementById(\'ct_date_training_passing\').value.trim(),\n    otet:document.getElementById(\'ct_otet\').value,\n    date_otet:document.getElementById(\'ct_date_otet\').value.trim(),\n    category:document.getElementById(\'ct_category\').value,\n    date_first_appt_ss:document.getElementById(\'ct_date_first_appt_ss\').value.trim(),\n    date_regular_teacher:document.getElementById(\'ct_date_regular_teacher\').value.trim(),\n    date_regularisation_consideration:document.getElementById(\'ct_date_regularisation_consideration\').value.trim(),\n    notification_no:document.getElementById(\'ct_notification_no\').value.trim(),\n    rank:document.getElementById(\'ct_rank\').value.trim(),\n    date_joining_levelVB:document.getElementById(\'ct_date_joining_levelVB\').value.trim(),\n    date_joining_levelIII:document.getElementById(\'ct_date_joining_levelIII\').value.trim(),\n    date_joining_levelIV:document.getElementById(\'ct_date_joining_levelIV\').value.trim(),\n    date_superannuation:document.getElementById(\'ct_date_superannuation\').value.trim(),\n    idt_ra:document.getElementById(\'ct_idt_ra\').value.trim(),\n    date_joining_district:document.getElementById(\'ct_date_joining_district\').value.trim(),\n    dp_cp_vigilance:document.getElementById(\'ct_dp_cp_vigilance\').value,\n    option:document.getElementById(\'ct_option\').value.trim(),\n    remark:document.getElementById(\'ct_remark\').value.trim()\n  };\n\n  const btn=document.getElementById(\'createBtn\');\n  btn.disabled=true;btn.textContent=\'⏳ Creating…\';\n  try{\n    const r=await fetch(\'/api/admin/teacher/create\',{\n      method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify(record)\n    });\n    const d=await r.json();\n    if(d.success){\n      showToast(\'Teacher created successfully!\',\'success\');\n      clearCreateForm();\n      await loadTeachers();await loadStats();await loadSchools();await loadDuplicates();\n      showTab(\'teachers\');\n    }else showToast(d.error||\'Creation failed\',\'error\');\n  }catch(e){showToast(\'Network error\',\'error\');}\n  finally{btn.disabled=false;btn.textContent=\'✅ Create Teacher Record\';}\n}\n\nfunction doExport(filter,status){\n  let url=\'/api/admin/export?type=xlsx\';\n  if(filter&&filter!==\'all\') url+=`&filter=${encodeURIComponent(filter)}`;\n  if(status&&status!==\'all\') url+=`&status=${encodeURIComponent(status)}`;\n  window.open(url,\'_blank\');\n}\n\n\nasync function loadDuplicates(){\n  try{\n    const r=await fetch(\'/api/admin/duplicates\');\n    const d=await r.json();\n    duplicateGroups=d.groups||[];\n    renderDuplicates();\n    const dupEl = document.getElementById(\'statDuplicates\'); if(dupEl) dupEl.textContent = d.group_count||0;\n  }catch(e){console.error(e);}\n}\n\nfunction renderDuplicates(){\n  const tbody=document.getElementById(\'duplicateTableBody\');\n  const info=document.getElementById(\'duplicateCount\');\n  if(!tbody) return;\n  tbody.innerHTML=\'\';\n  let rowCount=0;\n  if(!duplicateGroups.length){\n    tbody.innerHTML=\'<tr><td colspan="7" style="text-align:center;color:#999;padding:24px">No duplicate teachers across proformas found.</td></tr>\';\n    if(info) info.textContent=\'0 duplicate groups\';\n    return;\n  }\n  duplicateGroups.forEach((g, idx)=>{\n    (g.records||[]).forEach((rec, rix)=>{\n      rowCount++;\n      tbody.innerHTML+=`<tr>\n        <td>${rix===0?`<strong>#${idx+1}</strong><div style="font-size:.74rem;color:#777">${(g.proformas||[]).join(\', \')}</div>`:\'\'}</td>\n        <td><strong>${esc(rec.teacher_name)}</strong></td>\n        <td>${esc(rec.school||\'\')}</td>\n        <td>${esc(rec.dob||\'\')}</td>\n        <td>${esc(rec.proforma||\'\')}</td>\n        <td><span class="badge badge-${rec.status||\'PENDING\'}">${rec.status||\'PENDING\'}</span></td>\n        <td style="text-align:center;white-space:nowrap">\n          <button class="btn btn-outline btn-sm" onclick="openEditTeacher(\'${esc(rec.teacher_name)}\',\'${esc(rec.proforma)}\')">✏️ Edit</button>\n          <button class="btn btn-danger btn-sm" style="margin-left:4px" onclick="confirmDeleteTeacher(\'${esc(rec.teacher_name)}\',\'${esc(rec.proforma)}\')">🗑️</button>\n        </td>\n      </tr>`;\n    });\n  });\n  if(info) info.textContent=`${duplicateGroups.length} duplicate group(s), ${rowCount} record(s)`;\n}\n\n\nasync function submitBulkImport(){\n  const fileInput=document.getElementById(\'bulkFile\');\n  const proforma=document.getElementById(\'imp_proforma\').value;\n  const defaultSchool=document.getElementById(\'imp_default_school\').value;\n  const result=document.getElementById(\'importResult\');\n  if(!fileInput.files || !fileInput.files.length){showToast(\'Please choose an Excel file\',\'error\');return;}\n  if(!proforma){showToast(\'Please select a target proforma\',\'error\');return;}\n  const btn=document.getElementById(\'importBtn\');\n  btn.disabled=true; btn.textContent=\'⏳ Importing…\';\n  result.style.display=\'none\';\n  try{\n    const fd=new FormData();\n    fd.append(\'file\', fileInput.files[0]);\n    fd.append(\'proforma\', proforma);\n    fd.append(\'default_school\', defaultSchool);\n    const r=await fetch(\'/api/admin/import-excel\',{method:\'POST\',body:fd});\n    const d=await r.json();\n    if(d.success){\n      result.style.display=\'block\';\n      const warnings=(d.warnings||[]).slice(0,10).map(w=>`<li>${esc(w)}</li>`).join(\'\');\n      result.innerHTML=`\n        <div style="font-weight:700;color:#0d2366;margin-bottom:8px">✅ Import completed</div>\n        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin-bottom:12px">\n          <div style="background:#fff;border:1px solid #dde3f0;border-radius:8px;padding:10px"><div style="font-size:1.25rem;font-weight:700;color:#2e7d32">${d.created||0}</div><div style="font-size:.78rem;color:#666">Created</div></div>\n          <div style="background:#fff;border:1px solid #dde3f0;border-radius:8px;padding:10px"><div style="font-size:1.25rem;font-weight:700;color:#1565c0">${d.updated||0}</div><div style="font-size:.78rem;color:#666">Updated</div></div>\n          <div style="background:#fff;border:1px solid #dde3f0;border-radius:8px;padding:10px"><div style="font-size:1.25rem;font-weight:700;color:#e65100">${d.skipped||0}</div><div style="font-size:.78rem;color:#666">Skipped</div></div>\n          <div style="background:#fff;border:1px solid #dde3f0;border-radius:8px;padding:10px"><div style="font-size:1.25rem;font-weight:700;color:#6a1b9a">${d.processed_rows||0}</div><div style="font-size:.78rem;color:#666">Processed Rows</div></div>\n        </div>\n        <div style="font-size:.82rem;color:#555;margin-bottom:6px"><strong>Header row:</strong> ${d.header_row||\'-\'} &nbsp; <strong>Mapped fields:</strong> ${(d.mapped_fields||[]).join(\', \') || \'-\'}</div>\n        ${warnings?`<div style="margin-top:10px"><div style="font-size:.82rem;font-weight:700;color:#a15c00;margin-bottom:6px">Warnings</div><ul style="padding-left:18px;font-size:.82rem;color:#555">${warnings}</ul></div>`:\'\'}`;\n      showToast(`Import done: ${d.created||0} created, ${d.updated||0} updated`, \'success\');\n      await loadTeachers(); await loadStats(); await loadSchools(); await loadDuplicates();\n      populateSchoolDropdowns();\n    } else {\n      showToast(d.error||\'Import failed\',\'error\');\n      result.style.display=\'block\';\n      result.innerHTML=`<div style="font-weight:700;color:#c62828;margin-bottom:6px">Import failed</div><div style="font-size:.84rem;color:#555">${esc(d.error||\'Unknown error\')}</div>`;\n    }\n  }catch(e){\n    showToast(\'Network error during import\',\'error\');\n  }finally{\n    btn.disabled=false; btn.textContent=\'📤 Import Excel\';\n  }\n}\n\nasync function loadLogs(){\n  try{\n    const r=await fetch(\'/api/admin/logs?limit=100\');\n    const d=await r.json();\n    const logs=d.logs||[];\n    const wrap=document.getElementById(\'logContainer\');\n    if(!logs.length){wrap.innerHTML=\'<div style="padding:20px;text-align:center;color:#999">No activity yet.</div>\';return;}\n    wrap.innerHTML=logs.map(l=>`<div class="log-entry"><div class="log-time">${l.time||\'\'}</div><div class="log-msg">${esc(l.action||l.message||JSON.stringify(l))}</div></div>`).join(\'\');\n  }catch(e){console.error(e);}\n}\n\nfunction openModal(id){document.getElementById(id).classList.add(\'show\');}\nfunction closeModal(id){document.getElementById(id).classList.remove(\'show\');}\ndocument.querySelectorAll(\'.modal-overlay\').forEach(o=>{\n  o.addEventListener(\'click\',e=>{if(e.target===o) closeModal(o.id);});\n});\n\nfunction showToast(msg,type=\'info\'){\n  const t=document.getElementById(\'toast\');\n  t.textContent=msg;t.className=`toast ${type}`;t.classList.add(\'show\');\n  setTimeout(()=>t.classList.remove(\'show\'),3500);\n}\nfunction esc(s){if(s==null)return \'\';return String(s).replace(/&/g,\'&amp;\').replace(/</g,\'&lt;\').replace(/>/g,\'&gt;\').replace(/"/g,\'&quot;\');}\n</script>\n</body>\n</html>'

# ══ Seed Data ══

SEED_DATA = [{'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Balakrishna Savara',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': '+2 Science',
  'date_general_result': '21/06/2006',
  'training_qualification': 'D. El. Ed',
  'date_training_result': '20/02/2015',
  'otet': 'No',
  'date_otet': 'No',
  'category': 'ST',
  'dob': '30/04/1986',
  'date_first_appt_ss': '08/06/2011',
  'date_regular_teacher': '08/06/2017',
  'notification_no': '587/SME, Date-10/01/2011',
  'rank': '',
  'date_superannuation': '30/04/2046',
  'dp_cp_vigilance': 'NO',
  'idt_ra': '',
  'date_joining_district': '',
  'remark': 'Verified on 22/03/2026',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '2',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Nilambara Behera',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': 'BA',
  'date_general_result': '12/02/2021',
  'training_qualification': 'D. El. Ed',
  'date_training_result': '20/02/2015',
  'otet': 'OTET I',
  'date_otet': '20/10/2014',
  'category': 'SEBC',
  'dob': '05/10/1994',
  'date_first_appt_ss': '04/04/2018',
  'date_regular_teacher': '04/04/2024',
  'notification_no': '25611/SME, Date-26/12/2016',
  'rank': '138',
  'date_superannuation': '31/05/2054',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'IDT',
  'date_joining_district': '01/06/2026',
  'remark': 'Transferred from Berhampur on 01/03/2026',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '3',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Prakash Chandra Behera',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': '+2 Science',
  'date_general_result': '15/11/1994',
  'training_qualification': 'CT',
  'date_training_result': '31/12/1997',
  'otet': 'No',
  'date_otet': 'No',
  'category': 'SC',
  'dob': '05/04/1976',
  'date_first_appt_ss': '08/12/2003',
  'date_regular_teacher': '10/01/2009',
  'notification_no': '3479, Date-22/03/2003',
  'rank': '19',
  'date_superannuation': '31/05/2036',
  'dp_cp_vigilance': 'NO',
  'idt_ra': '',
  'date_joining_district': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '4',
  'district': 'CHIKITI',
  'teacher_name': 'BIJAYALAXMI PADHY',
  'school': 'GOVT. PS. NEW MAISANPUR',
  'general_qualification': 'I. Sc.',
  'date_general_result': '30.05.2011',
  'training_qualification': 'C.T',
  'date_training_result': '07/09/2013',
  'otet': 'I',
  'date_otet': '03.07.2013',
  'category': 'UR',
  'dob': '20/05/1994',
  'date_first_appt_ss': '10.12.2013',
  'date_regular_teacher': '01/01/2020',
  'notification_no': '',
  'rank': '',
  'date_superannuation': '31.05.2054',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '5',
  'district': 'CHIKITI',
  'teacher_name': 'SUBRAT KUMAR PRADHAN',
  'school': 'GOVT. PS. NEW MAISANPUR',
  'general_qualification': 'I. Sc.',
  'date_general_result': '26.05.2009',
  'training_qualification': 'C.T',
  'date_training_result': '31.12.2011',
  'otet': 'I',
  'date_otet': '02.01.2013',
  'category': 'UR',
  'dob': '10.04.1992',
  'date_first_appt_ss': '30.11.2012',
  'date_regular_teacher': '30.11.2018',
  'notification_no': '',
  'rank': '',
  'date_superannuation': '30.04.2052',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '6',
  'district': 'CHIKITI',
  'teacher_name': 'P. MEENAKETANA SAHU',
  'school': 'GPS, PARASAMBA',
  'general_qualification': 'I.Com',
  'date_general_result': '18.07.1998',
  'training_qualification': 'C.T',
  'date_training_result': '21.12.200',
  'otet': 'I',
  'date_otet': '15.11.2025',
  'category': 'UR',
  'dob': '01.06.1980',
  'date_first_appt_ss': '10.03.2005',
  'date_regular_teacher': '12.04.2011',
  'notification_no': 'N A',
  'rank': '14/BME/CT/UR/M-',
  'date_superannuation': '31.05.2040',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '7',
  'district': 'CHIKITI',
  'teacher_name': 'TANUJA SABAT',
  'school': 'GPS, PARASAMBA',
  'general_qualification': 'I. SC',
  'date_general_result': '30.05.2011',
  'training_qualification': 'C.T',
  'date_training_result': '07.09.2013',
  'otet': 'I',
  'date_otet': '01.09.2014',
  'category': 'UR',
  'dob': '17.05.1994',
  'date_first_appt_ss': '19.01.2015',
  'date_regular_teacher': '19.01.2021',
  'notification_no': '7610/11..09.2014',
  'rank': '35',
  'date_superannuation': '31.05.2054',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '8',
  'district': 'CHIKITI',
  'teacher_name': 'KAMSU ANKATI RAO',
  'school': 'GPS. K. K. PUR',
  'general_qualification': 'B.A',
  'date_general_result': '30/06/1993',
  'training_qualification': 'BED',
  'date_training_result': '03/12/1998',
  'otet': 'NO',
  'date_otet': 'NO',
  'category': 'UR',
  'dob': '02/10/1971',
  'date_first_appt_ss': '13/12/2006',
  'date_regular_teacher': '21/12/2012',
  'notification_no': 'NA',
  'rank': 'NA',
  'date_superannuation': '28/02/2031',
  'dp_cp_vigilance': 'NA',
  'idt_ra': 'NA',
  'date_joining_district': 'NA',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '9',
  'district': 'CHIKITI',
  'teacher_name': 'B SILPA KUMARI PATRO',
  'school': 'GUPS MAISANPUR',
  'general_qualification': 'PLUS 2 ARTS',
  'date_general_result': '29/05/2013',
  'training_qualification': 'CT',
  'date_training_result': '05/06/2016',
  'otet': 'OTET -1',
  'date_otet': '12/01/2015',
  'category': 'UR',
  'dob': '26/05/1994',
  'date_first_appt_ss': '04/02/2018',
  'date_regular_teacher': '04/02/2024',
  'notification_no': '12511/SME26/12/2016',
  'rank': 'ARTS UR-51',
  'date_superannuation': '31/05/2054',
  'dp_cp_vigilance': '',
  'idt_ra': '',
  'date_joining_district': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '10',
  'district': 'CHIKITI',
  'teacher_name': 'SUNIL KUMAR PRADHAN',
  'school': 'GUPS, CHILADI',
  'general_qualification': 'I. Sc.',
  'date_general_result': '26.05.2009',
  'training_qualification': 'C.T',
  'date_training_result': '31.12.2011',
  'otet': 'I',
  'date_otet': '03.07.2013',
  'category': 'UR',
  'dob': '15.04.1992',
  'date_first_appt_ss': '30.11.2012',
  'date_regular_teacher': '30.11.2018',
  'notification_no': 'N A',
  'rank': '',
  'date_superannuation': '30.04.2052',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '11',
  'district': 'CHIKITI',
  'teacher_name': 'MALLISWARI DOKKARI',
  'school': 'SRI RADHAKANTA NODAL HIGH SCHOOL KOTILINGI',
  'general_qualification': 'PLUS TWO ARTS',
  'date_general_result': '28.04.2011',
  'training_qualification': 'CT',
  'date_training_result': '21/12/2013',
  'otet': 'OTET P-1',
  'date_otet': '09/01/2014',
  'category': 'UR',
  'dob': '25/04/1987',
  'date_first_appt_ss': '19/01/2015',
  'date_regular_teacher': '19/01/2021',
  'notification_no': '20406 SME DATE 11/09/2014',
  'rank': '',
  'date_superannuation': '30/04/2047',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '12',
  'district': 'GANJAM/CHIKITI',
  'teacher_name': 'TULASIRAO GUMMIDI',
  'school': 'SRK NODAL HIGH SCHOOL KOTILINGI',
  'general_qualification': 'SCIENCE',
  'date_general_result': '06.08.2007',
  'training_qualification': 'B.Ed',
  'date_training_result': '26.10.2009',
  'otet': 'II',
  'date_otet': '11.03.2015',
  'category': 'UR',
  'dob': '04.03.1987',
  'date_first_appt_ss': '30.11.2012',
  'date_regular_teacher': '30.11.2018',
  'notification_no': '5458  (08.12.2011)',
  'rank': '40',
  'date_superannuation': '31.03.2047',
  'dp_cp_vigilance': 'NA',
  'idt_ra': 'NA',
  'date_joining_district': 'NA',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA I',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '13',
  'district': 'CHIKITI',
  'teacher_name': 'BARU PUJA',
  'school': 'UPS CHILADI',
  'general_qualification': 'I. A',
  'date_general_result': '30.05.2011',
  'training_qualification': 'C.T',
  'date_training_result': '07.09.2013',
  'otet': 'I',
  'date_otet': '03.07.2013',
  'category': 'SC',
  'dob': '22.10.1993',
  'date_first_appt_ss': '10.12.2013',
  'date_regular_teacher': '10.12.2019',
  'notification_no': 'N A',
  'rank': '',
  'date_superannuation': '31.10.2053',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Jyoti Maharana',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': 'Graduation BA',
  'date_general_result': '16/04/2011',
  'training_qualification': 'B. Ed',
  'date_training_result': '07/03/2013',
  'otet': 'OTET II',
  'date_otet': '09/01/2014',
  'category': 'UR',
  'dob': '04/08/1989',
  'date_first_appt_ss': '19/01/2015',
  'date_regular_teacher': '19/01/2021',
  'notification_no': '20406/SME, Date-11/09/2014',
  'rank': '23',
  'date_superannuation': '30/04/2049',
  'dp_cp_vigilance': 'NO',
  'idt_ra': '',
  'date_joining_district': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '2',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Nilambara Behera',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': 'BA',
  'date_general_result': '12/02/2021',
  'training_qualification': 'D. El. Ed',
  'date_training_result': '20/02/2015',
  'otet': 'OTET II',
  'date_otet': '22/11/2024',
  'category': 'SEBC',
  'dob': '05/10/1994',
  'date_first_appt_ss': '04/04/2018',
  'date_regular_teacher': '04/04/2024',
  'notification_no': '25611/SME, Date-26/12/2016',
  'rank': '138',
  'date_superannuation': '31/05/2054',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'IDT',
  'date_joining_district': '01/06/2026',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '3',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Radharani Patra',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': 'B. Sc',
  'date_general_result': '19/05/2014',
  'training_qualification': 'CT',
  'date_training_result': '24/12/2010',
  'otet': 'OTET II',
  'date_otet': '15/11/2025',
  'category': 'UR',
  'dob': '16/06/1991',
  'date_first_appt_ss': '13/05/2011',
  'date_regular_teacher': '13/05/2017',
  'notification_no': '587/SME, Date-10/01/2011',
  'rank': '1023',
  'date_superannuation': '30/06/2051',
  'dp_cp_vigilance': 'NO',
  'idt_ra': '',
  'date_joining_district': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '4',
  'district': 'CHIKITI',
  'teacher_name': 'P. MEENAKETANA SAHU',
  'school': 'GPS, PARASAMBA',
  'general_qualification': 'B. COM',
  'date_general_result': '14.10.2003',
  'training_qualification': 'C.T',
  'date_training_result': '21.12.2000',
  'otet': 'II',
  'date_otet': '25.11.2025',
  'category': 'UR',
  'dob': '01.06.1980',
  'date_first_appt_ss': '10.03.2005',
  'date_regular_teacher': '12.04.2011',
  'notification_no': '15.08.2004',
  'rank': '14/BME/CT/UR M',
  'date_superannuation': '31.05.2040',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '5',
  'district': 'CHIKITI',
  'teacher_name': 'KAMSU ANKATI RAO',
  'school': 'GPS. K. K. PUR',
  'general_qualification': 'B. A',
  'date_general_result': '30/06/1993',
  'training_qualification': 'BED',
  'date_training_result': '03/12/1998',
  'otet': 'll',
  'date_otet': '15/11/2025',
  'category': 'UR',
  'dob': '02/10/1971',
  'date_first_appt_ss': '13/12/2006',
  'date_regular_teacher': '21/12/2012',
  'notification_no': 'NA',
  'rank': 'NA',
  'date_superannuation': '28/02/2031',
  'dp_cp_vigilance': 'NA',
  'idt_ra': 'NA',
  'date_joining_district': 'NA',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '6',
  'district': 'CHIKITI',
  'teacher_name': 'GOURI SANKAR SAHU',
  'school': 'GUPS BADABARANGA',
  'general_qualification': 'B.SC',
  'date_general_result': '01.07.2007',
  'training_qualification': 'B.ED',
  'date_training_result': '09.08.2010',
  'otet': 'II',
  'date_otet': '15.11.2025',
  'category': 'UR',
  'dob': '10.06.1987',
  'date_first_appt_ss': '13.05.2011',
  'date_regular_teacher': '13.05.2017',
  'notification_no': '587/10.01.2011',
  'rank': '',
  'date_superannuation': '30.06.2047',
  'dp_cp_vigilance': 'No',
  'idt_ra': 'No',
  'date_joining_district': 'no',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '7',
  'district': 'CHIKITI',
  'teacher_name': 'KAMESWARI PANIGRAHY',
  'school': 'GUPS BADABARANGA',
  'general_qualification': 'B.Sc',
  'date_general_result': '19.05.2014',
  'training_qualification': 'CT',
  'date_training_result': '06.01.2011',
  'otet': 'II',
  'date_otet': '15.11.2025',
  'category': 'UR',
  'dob': '29.05.1991',
  'date_first_appt_ss': '13.05.2011',
  'date_regular_teacher': '07.07.2017',
  'notification_no': '587/10.01.2011',
  'rank': '',
  'date_superannuation': '31.05.2051',
  'dp_cp_vigilance': 'No',
  'idt_ra': 'No',
  'date_joining_district': 'no',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '8',
  'district': 'CHIKITI',
  'teacher_name': 'NAMITA BEHERA',
  'school': 'GUPS BADABARANGA',
  'general_qualification': 'B.A',
  'date_general_result': '16/12/2019',
  'training_qualification': 'CT',
  'date_training_result': '17.11.2005',
  'otet': 'II',
  'date_otet': '15.11.2025',
  'category': 'SEBC',
  'dob': '12.10.1974',
  'date_first_appt_ss': '13.12.2006',
  'date_regular_teacher': '21.12.2012',
  'notification_no': '11676/SME/31.05.2006',
  'rank': '',
  'date_superannuation': '31.10.2034',
  'dp_cp_vigilance': 'No',
  'idt_ra': 'No',
  'date_joining_district': 'no',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '9',
  'district': 'CHIKITI',
  'teacher_name': 'BARU PUJA',
  'school': 'GUPS, CHILADI',
  'general_qualification': 'B. A',
  'date_general_result': '11.05.2017',
  'training_qualification': 'C.T',
  'date_training_result': '07.09.2013',
  'otet': 'II',
  'date_otet': '15.11.2025',
  'category': 'SC',
  'dob': '22.10.1993',
  'date_first_appt_ss': '10.12.2013',
  'date_regular_teacher': '10.12.2019',
  'notification_no': 'N A',
  'rank': '',
  'date_superannuation': '31.10.2053',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '10',
  'district': 'CHIKITI',
  'teacher_name': 'MANINI SAHU',
  'school': 'GUPS, CHILADI',
  'general_qualification': 'B. A',
  'date_general_result': '30.08.2010',
  'training_qualification': 'B. Ed',
  'date_training_result': '19.01.2014',
  'otet': 'II',
  'date_otet': '03.07.2013',
  'category': 'UR',
  'dob': '07.05.1989',
  'date_first_appt_ss': '01.08.2015',
  'date_regular_teacher': '01.08.2021',
  'notification_no': 'N A',
  'rank': '',
  'date_superannuation': '31.05.2049',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '11',
  'district': 'CHIKITI',
  'teacher_name': 'SUNIL KUMAR PRADHAN',
  'school': 'GUPS, CHILADI',
  'general_qualification': 'B. SC',
  'date_general_result': '09.11.2015',
  'training_qualification': 'C.T',
  'date_training_result': '31.12.2011',
  'otet': 'II',
  'date_otet': '22.11.2024',
  'category': 'UR',
  'dob': '15.04.1992',
  'date_first_appt_ss': '30.11.2012',
  'date_regular_teacher': '30.11.2018',
  'notification_no': 'N A',
  'rank': '',
  'date_superannuation': '30.04.2052',
  'dp_cp_vigilance': 'N A',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '12',
  'district': 'CHIKITI',
  'teacher_name': 'NEELAPU KESAVA RAO',
  'school': 'SRI RADHA KANTA HIGH SCHOOL,. KOTILINGI',
  'general_qualification': 'B. A',
  'date_general_result': '07/02/2004',
  'training_qualification': 'B. Ed',
  'date_training_result': '05/07/2006',
  'otet': 'OTET2',
  'date_otet': '15-11-2025',
  'category': 'SEBC',
  'dob': '07/01/1978',
  'date_first_appt_ss': '13-12-2012',
  'date_regular_teacher': '21-12-2012',
  'notification_no': '11676 SME Dt.31/05/2006',
  'rank': '',
  'date_superannuation': '30-06-2038',
  'dp_cp_vigilance': 'NO',
  'idt_ra': '',
  'date_joining_district': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '13',
  'district': 'GANJAM',
  'teacher_name': 'DILLESWAR RAULO',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'BA',
  'date_general_result': '25.07.2012',
  'training_qualification': 'CT',
  'date_training_result': '20.06.2010',
  'otet': 'OTET P-2',
  'date_otet': '15-11-2025',
  'category': 'UR',
  'dob': '03.05.1990',
  'date_first_appt_ss': '21.12.2011',
  'date_regular_teacher': '21.12.2017',
  'notification_no': '587/10.01.2011',
  'rank': '',
  'date_superannuation': '31.05.2050',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '14',
  'district': 'CHIKITI',
  'teacher_name': 'JAGANNAIKULU KORRAYI',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'Bsc',
  'date_general_result': '28/07/2007',
  'training_qualification': 'BED',
  'date_training_result': '27/11/2008',
  'otet': 'P-2',
  'date_otet': '03/11/2015',
  'category': 'UR',
  'dob': '07/09/1986',
  'date_first_appt_ss': '30/11/2012',
  'date_regular_teacher': '30/12/2018',
  'notification_no': '5458 SME 08/12/2011',
  'rank': '',
  'date_superannuation': '31/07/2046',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '15',
  'district': 'GANJAM',
  'teacher_name': 'KEDARNATH RAULO',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'BA',
  'date_general_result': '11.05.2017',
  'training_qualification': 'CT',
  'date_training_result': '28.02.2009',
  'otet': 'OTET P-2',
  'date_otet': '15-11-2025',
  'category': 'UR',
  'dob': '08.05.1988',
  'date_first_appt_ss': '13.05.2011',
  'date_regular_teacher': '13.05.2017',
  'notification_no': '587/10.01.2011',
  'rank': '',
  'date_superannuation': '31.05.2048',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '16',
  'district': 'CHIKITI',
  'teacher_name': 'KORAI UMAMAHESHWAR RAO',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'BCOM',
  'date_general_result': '28.08.1998',
  'training_qualification': 'CT',
  'date_training_result': '02.06.2006',
  'otet': 'OTET P-2',
  'date_otet': '15/11/2025',
  'category': 'UR',
  'dob': '05/04/1979',
  'date_first_appt_ss': '13/12/2006',
  'date_regular_teacher': '21/12/2012',
  'notification_no': '11676 SME DATE 31/05/2006',
  'rank': '',
  'date_superannuation': '31/05/2039',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA II',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '17',
  'district': 'CHIKITI',
  'teacher_name': 'KARINGI TULASAYYA',
  'school': 'SRI RADHAKANTA NODAL HIGH SCHOOL KOTILINGI',
  'general_qualification': 'BA',
  'date_general_result': '04.07.1998',
  'training_qualification': 'CT',
  'date_training_result': '02.06.2006',
  'otet': 'PAPER 2',
  'date_otet': '15/11/2025',
  'category': 'UR',
  'dob': '17/02/1975',
  'date_first_appt_ss': '13/12/2006',
  'date_regular_teacher': '21/12/2012',
  'notification_no': '11676 SME DATE 31/05/2006',
  'rank': '',
  'date_superannuation': '28/02/2035',
  'dp_cp_vigilance': 'NO',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA III',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Pitambar Sethi',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': '+2 Arts CT',
  'date_general_result': '08/06/2007',
  'training_qualification': 'CT',
  'date_training_result': '17/11/2005',
  'otet': 'NA',
  'date_otet': 'NA',
  'category': 'SC',
  'dob': '25/05/1971',
  'date_first_appt_ss': '16/12/1991',
  'date_regular_teacher': '01/12/1994',
  'notification_no': 'Sikshya Karmi , 8742/91, Date- 29/11/1991',
  'rank': '60',
  'date_joining_levelVB': '30/09/2020',
  'date_superannuation': '31/05/2031',
  'idt_ra': '',
  'date_joining_district': 'NA',
  'dp_cp_vigilance': 'No',
  'option': 'Not Necessary',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA III',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '2',
  'district': 'CHIKITI',
  'teacher_name': 'PRAMOD KUMAR PADHY',
  'school': 'GUPS BADABARANGA',
  'general_qualification': 'matric',
  'date_general_result': '15.12.1991',
  'training_qualification': 'CT',
  'date_training_result': '31.12.2001',
  'otet': 'no',
  'date_otet': 'not appeared',
  'category': 'UR',
  'dob': '05.06.1971',
  'date_first_appt_ss': '',
  'date_regular_teacher': '18.08.1994',
  'notification_no': '',
  'rank': '',
  'date_joining_levelVB': '30.09.2020',
  'date_superannuation': '30.06.2031',
  'idt_ra': 'no',
  'date_joining_district': 'no',
  'dp_cp_vigilance': 'no',
  'option': '',
  'remark': 'fergue of level 4 promotion',
  'change_history': []},
 {'proforma': 'PROFORMA IV',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'ganjam, chikiti',
  'teacher_name': 'RAMAKANT SAHU',
  'school': 'GUPS MAISANPUR',
  'general_qualification': 'plus 2 sc',
  'date_general_result': '1993',
  'training_qualification': 'CT',
  'date_training_result': '24/11/1997',
  'otet': 'OTET I',
  'date_otet': '15/11/2025',
  'category': 'UR',
  'dob': '24/06/1976',
  'date_first_appt_ss': '20/10/2001',
  'date_regular_teacher': '01/04/2008',
  'notification_no': '14/10/2000',
  'rank': '',
  'date_joining_levelVB': '08/01/2022',
  'date_superannuation': '30/06/1936',
  'idt_ra': 'NO',
  'date_joining_district': '',
  'dp_cp_vigilance': '',
  'option': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA V',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'CHIKITI',
  'teacher_name': 'BAIRAGI BEHJERA',
  'school': 'GUPS, CHILADI',
  'general_qualification': 'I. Sc',
  'date_general_result': '01/05/1994',
  'training_qualification': 'C.T.',
  'date_training_result': '31.12.1997',
  'otet': 'II',
  'date_otet': '15.11.2025',
  'category': 'UR',
  'dob': '18.06.1976',
  'date_first_appt_ss': '08.08.2003',
  'date_regular_teacher': '27.09.2009',
  'notification_no': 'N A',
  'rank': '',
  'date_joining_levelVB': '10.01.2022',
  'date_superannuation': '30.06.2036',
  'idt_ra': 'NO',
  'date_joining_district': '',
  'dp_cp_vigilance': 'N. A',
  'option': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA V',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '2',
  'district': 'CHIKITI',
  'teacher_name': 'G KRISHNAMURTY',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'BA',
  'date_general_result': '27/06/1996',
  'training_qualification': 'BED',
  'date_training_result': '24.12.1999',
  'otet': 'OTET P=2',
  'date_otet': '12.02.2026',
  'category': 'UR',
  'dob': '06/04/1975',
  'date_first_appt_ss': '25/07/2003',
  'date_regular_teacher': '13/09/2009',
  'notification_no': '3479 SME 22/03/2003',
  'rank': '',
  'date_joining_levelVB': '01/08/2022',
  'date_superannuation': '30.06.2035',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'dp_cp_vigilance': 'NO',
  'option': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA V',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '3',
  'district': 'CHIKITI',
  'teacher_name': 'JAGANNATH PANDA',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'BA',
  'date_general_result': '22.07.2004',
  'training_qualification': 'BED',
  'date_training_result': '25.02.2014',
  'otet': 'OTET P=2',
  'date_otet': '12.02.2026',
  'category': 'UR',
  'dob': '11.04.1977',
  'date_first_appt_ss': '04.05.2001',
  'date_regular_teacher': '01.04.2008',
  'notification_no': '27021 SME 03.10.2000',
  'rank': '',
  'date_joining_levelVB': '08.01.2022',
  'date_superannuation': '30.04.2037',
  'idt_ra': 'NO',
  'date_joining_district': 'NO',
  'dp_cp_vigilance': 'NO',
  'option': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VI',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'Ganjam/Chikiti',
  'teacher_name': 'Mahesweta Patra',
  'school': 'Barahi Bidyapitha Kalabada',
  'general_qualification': 'BA',
  'date_passing': '30/06/2006',
  'training_qualification': 'B. Ed',
  'date_training_passing': '25/02/2014',
  'otet': 'OTET I & II',
  'date_otet': 'OTET I- 20/10/2014,                    OTET II- 15/11/2025',
  'category': 'ST',
  'dob': '04/02/1986',
  'date_first_appt_ss': '09/03/2005',
  'date_regular_teacher': '04/12/2011',
  'notification_no': '14281/SME, Date-19/08/2004',
  'rank': '',
  'date_joining_levelIV': '23/10/2021',
  'date_superannuation': '30/04/2046',
  'idt_ra': '',
  'date_joining_district': '',
  'dp_cp_vigilance': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VI',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '2',
  'district': 'CHIKITI',
  'teacher_name': 'RUSHI PRADHAN',
  'school': 'GUPS BADABARANGA',
  'general_qualification': 'B.A',
  'date_passing': '22.07.2004',
  'training_qualification': 'CT',
  'date_training_passing': '17.10.1992',
  'otet': 'NO',
  'date_otet': 'not appeared',
  'category': 'SEBC',
  'dob': '19.03.1968',
  'date_first_appt_ss': '',
  'date_regular_teacher': '30.11.1996',
  'notification_no': '',
  'rank': '',
  'date_joining_levelIV': '25.10.2021',
  'date_superannuation': '31.03.2028',
  'idt_ra': 'no',
  'date_joining_district': 'no',
  'dp_cp_vigilance': 'no',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VI',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '3',
  'district': 'Ganjm',
  'teacher_name': 'BHAGABAN PANIGRAHY',
  'school': 'GUPS MAISANPUR',
  'general_qualification': '+3 Arts',
  'date_passing': '19/2/2021',
  'training_qualification': 'CT',
  'date_training_passing': '31/07/1996',
  'otet': 'OTET P-2',
  'date_otet': '15/11/2025',
  'category': 'UR',
  'dob': '09/07/1974',
  'date_first_appt_ss': '',
  'date_regular_teacher': '29/11/1996',
  'notification_no': '',
  'rank': '31',
  'date_joining_levelIV': '22/10/2021',
  'date_superannuation': '30/09/2034',
  'idt_ra': '',
  'date_joining_district': '',
  'dp_cp_vigilance': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VI',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '4',
  'district': 'CHIKITI',
  'teacher_name': 'AJAY KUMAR PRADHAN',
  'school': 'GUPS, CHILADI',
  'general_qualification': 'B. A',
  'date_passing': '14.08.1991',
  'training_qualification': 'B. Ed',
  'date_training_passing': '29.09.1994',
  'otet': 'NO',
  'date_otet': 'N A',
  'category': 'UR',
  'dob': '05.06.1966',
  'date_first_appt_ss': 'N A',
  'date_regular_teacher': '27.11.1996',
  'notification_no': '',
  'rank': '',
  'date_joining_levelIV': '23.10.2021',
  'date_superannuation': '30.06.2026',
  'idt_ra': 'N A',
  'date_joining_district': 'N A',
  'dp_cp_vigilance': 'N A',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VI',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '5',
  'district': 'CHIKITI',
  'teacher_name': 'ALABALU NARASINGH',
  'school': 'SRI RADHAKANTA HIGH SCHOOL KOTILINGI',
  'general_qualification': 'B. A',
  'date_passing': '01.07.2007',
  'training_qualification': 'CT',
  'date_training_passing': '24/12/2010',
  'otet': 'OTET P=2 and PAPER =1',
  'date_otet': '15.11.2025 AND 12/02/2026',
  'category': 'ST',
  'dob': '12.07.1983',
  'date_first_appt_ss': '13.12.2006',
  'date_regular_teacher': '21.12.2012',
  'notification_no': '11676 SME DATE 31.05.2006',
  'rank': '',
  'date_joining_levelIV': '23.10.2021',
  'date_superannuation': '31.07.2043',
  'idt_ra': 'N0',
  'date_joining_district': 'No',
  'dp_cp_vigilance': 'NO',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VII',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '1',
  'district': 'CHIKITI',
  'teacher_name': 'B. MOHAN RAO',
  'school': 'GPS, PARASAMBA',
  'general_qualification': 'HSC',
  'date_passing': '31.01.1985',
  'training_qualification': 'CT',
  'date_training_passing': '18.12.2004',
  'otet': 'NO',
  'date_otet': 'N A',
  'category': 'SC',
  'dob': '14.02.1968',
  'date_first_appt_ss': '02.01.1992',
  'date_regular_teacher': '01.12.1994',
  'notification_no': '',
  'rank': '',
  'date_joining_levelIV': '22.10.2021',
  'date_superannuation': '28.02.2028',
  'idt_ra': '',
  'date_joining_district': '',
  'dp_cp_vigilance': '',
  'remark': '',
  'change_history': []},
 {'proforma': 'PROFORMA VII',
  'status': 'PENDING',
  'last_updated': '',
  'updated_by': '',
  'sl_no': '2',
  'district': 'CHIKITI',
  'teacher_name': 'DEVARAJ MALLIK',
  'school': 'GPS. K. K. PUR',
  'general_qualification': '+2 ARTS',
  'date_passing': '26/05/2009',
  'training_qualification': 'CT',
  'date_training_passing': '07/07/2012',
  'otet': 'I',
  'date_otet': '02/12/2026',
  'category': 'ST',
  'dob': '02/08/1984',
  'date_first_appt_ss': '03/09/2005',
  'date_regular_teacher': '04/11/2011',
  'notification_no': '15/08/2004',
  'rank': '278/BAM/UT/ST M',
  'date_joining_levelIV': '23/10/2021',
  'date_superannuation': '28/02/2044',
  'idt_ra': 'NA',
  'date_joining_district': 'NA',
  'dp_cp_vigilance': 'NA',
  'remark': '',
  'change_history': []}]

# ══ Field Config ══

FIELD_LABELS = {'sl_no': 'S.No.',
 'district': 'District / Block',
 'teacher_name': 'Name of Teacher',
 'school': 'School Name',
 'general_qualification': 'General Qualification',
 'date_general_result': 'Date of Result (General)',
 'date_passing': 'Date of Passing (General)',
 'training_qualification': 'Training Qualification',
 'date_training_result': 'Date of Result (Training)',
 'date_training_passing': 'Date of Passing (Training)',
 'otet': 'OTET (I / II / NO)',
 'date_otet': 'Date of Passing OTET',
 'category': 'Category',
 'dob': 'Date of Birth',
 'date_first_appt_ss': 'Date of 1st Appointment (SS)',
 'date_regular_teacher': 'Date of Regular Appointment',
 'notification_no': 'Notification No. & Date',
 'rank': 'Rank in Merit List',
 'date_joining_levelVB': 'Date of Joining Level-V(B)',
 'date_joining_levelIV': 'Date of Joining Level-IV',
 'date_superannuation': 'Date of Superannuation',
 'idt_ra': 'IDT / RA',
 'date_joining_district': 'Date of Joining District',
 'dp_cp_vigilance': 'DP/CP/Vigilance Pending?',
 'option': 'Option',
 'remark': 'Remark',
 'date_joining_levelIII': 'Date of Joining in Level-III Post',
 'date_regularisation_consideration': 'Date of Regularisation for Consideration'}

FIELD_GROUPS = [{'title': '🏫 School & Location', 'fields': ['district', 'school']},
 {'title': '🎓 Qualifications',
  'fields': ['general_qualification',
             'date_general_result',
             'date_passing',
             'training_qualification',
             'date_training_result',
             'date_training_passing',
             'otet',
             'date_otet']},
 {'title': '📋 Personal Details', 'fields': ['category', 'dob']},
 {'title': '📅 Service Details',
  'fields': ['date_first_appt_ss',
             'date_regular_teacher',
             'date_regularisation_consideration',
             'notification_no',
             'rank',
             'date_joining_levelVB',
             'date_joining_levelIII',
             'date_joining_levelIV',
             'date_superannuation',
             'idt_ra',
             'date_joining_district',
             'dp_cp_vigilance',
             'option',
             'remark']}]

PROTECTED_FIELDS = {'sl_no','teacher_name','proforma','status','last_updated','updated_by','change_history'}

# ══ Backend ══

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DATA_PATH = os.path.join(DATA_DIR, "teachers.json")
SCHOOLS_PATH = os.path.join(DATA_DIR, "schools.json")
LOG_PATH = os.path.join(DATA_DIR, "activity_log.json")


def ensure_data_files():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(DATA_PATH) or os.path.getsize(DATA_PATH) < 10:
        with open(DATA_PATH, "w") as f:
            json.dump(SEED_DATA, f, indent=2)
    if not os.path.exists(SCHOOLS_PATH):
        schools = sorted(set(t["school"] for t in SEED_DATA if t.get("school")))
        with open(SCHOOLS_PATH, "w") as f:
            json.dump(schools, f, indent=2)
    if not os.path.exists(LOG_PATH):
        with open(LOG_PATH, "w") as f:
            json.dump([], f)


ensure_data_files()


def load_teachers():
    try:
        with open(DATA_PATH) as f:
            return json.load(f)
    except Exception:
        return list(SEED_DATA)


def save_teachers(teachers):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DATA_PATH, "w") as f:
        json.dump(teachers, f, indent=2)


def load_schools():
    schools = set()
    try:
        with open(SCHOOLS_PATH) as f:
            for s in json.load(f):
                if s:
                    schools.add(s.strip())
    except Exception:
        pass
    for t in load_teachers():
        if t.get("school"):
            schools.add(t["school"].strip())
    return sorted(schools)


def save_schools(schools):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(SCHOOLS_PATH, "w") as f:
        json.dump(sorted(set(s.strip() for s in schools if s.strip())), f, indent=2)


def get_schools():
    return load_schools()


def normalize_dob(dob_str):
    if not dob_str:
        return ""
    s = str(dob_str).strip().replace(".", "/").replace("-", "/")
    s = re.sub(r"\s+", "", s)
    return s.upper()


def dob_matches(stored_dob, entered_dob):
    return normalize_dob(stored_dob) == normalize_dob(entered_dob)


def normalize_teacher_key(name):
    return re.sub(r'\s+', ' ', (name or '').strip().upper())


def find_duplicate_groups(teachers):
    grouped = {}
    for t in teachers:
        key = normalize_teacher_key(t.get('teacher_name', ''))
        if not key:
            continue
        grouped.setdefault(key, []).append(t)
    results = []
    for key, rows in grouped.items():
        proformas = sorted({r.get('proforma', '') for r in rows if r.get('proforma')})
        if len(proformas) > 1:
            results.append({
                'teacher_name': rows[0].get('teacher_name', ''),
                'normalized_name': key,
                'proformas': proformas,
                'records': rows
            })
    results.sort(key=lambda x: (x['teacher_name'] or '').upper())
    return results


def normalize_header_name(value):
    s = str(value or '').strip().lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    return s


IMPORT_HEADER_ALIASES = {
    'sl no': 'sl_no', 'sl no ': 'sl_no', 'sl no.': 'sl_no', 'sl no': 'sl_no', 'sl': 'sl_no', 'serial no': 'sl_no', 'serial number': 'sl_no',
    'district': 'district', 'district block': 'district', 'district block name': 'district', 'district block ': 'district',
    'name of teacher': 'teacher_name', 'teacher name': 'teacher_name', 'name': 'teacher_name', 'name of the ups hm': 'teacher_name', 'name of ups hm': 'teacher_name',
    'school': 'school', 'school name': 'school',
    'general qualification': 'general_qualification',
    'date of result general': 'date_general_result', 'date general result': 'date_general_result', 'date of publication of b a b sc result': 'date_general_result', 'date of publication of ba b sc result': 'date_general_result',
    'date of passing general': 'date_passing', 'date passing general': 'date_passing',
    'training qualification': 'training_qualification',
    'date of result training': 'date_training_result', 'date training result': 'date_training_result', 'date of publication of ct b ed result': 'date_training_result',
    'date of passing training': 'date_training_passing', 'date training passing': 'date_training_passing',
    'whether passed otet i ii': 'otet', 'whether passed otet i ii ': 'otet', 'otet': 'otet', 'whether passed otet': 'otet',
    'date of publication of otet i ii result': 'date_otet', 'date of passing otet': 'date_otet', 'date otet': 'date_otet',
    'category': 'category',
    'date of birth': 'dob', 'dob': 'dob',
    'date of 1st appointment ss': 'date_first_appt_ss', 'date of first appointment as regular teacher level v': 'date_first_appt_ss', 'date of first appointment ss': 'date_first_appt_ss',
    'date of regular appointment': 'date_regular_teacher',
    'date of regularisation for consideration': 'date_regularisation_consideration',
    'notification no date': 'notification_no', 'notification no': 'notification_no',
    'rank': 'rank', 'rank in merit list': 'rank',
    'date of joining level v b': 'date_joining_levelVB',
    'date of joining in level iii post': 'date_joining_levelIII', 'date of joining level iii': 'date_joining_levelIII',
    'date of joining level iv': 'date_joining_levelIV',
    'date of superannuation': 'date_superannuation',
    'idt ra': 'idt_ra',
    'date of joining district': 'date_joining_district',
    'dp cp vigilance pending': 'dp_cp_vigilance', 'dp cp vigilance pending ': 'dp_cp_vigilance',
    'option': 'option',
    'remark': 'remark',
    'status': 'status',
    'proforma': 'proforma'
}


def detect_header_row(ws, max_scan_rows=6):
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [normalize_header_name(c.value) for c in ws[row_idx] if c.value not in (None, '')]
        score = sum(1 for v in values if v in IMPORT_HEADER_ALIASES)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row if best_score >= 1 else 1


def build_import_column_map(ws, header_row):
    col_map = {}
    for cell in ws[header_row]:
        normalized = normalize_header_name(cell.value)
        field = IMPORT_HEADER_ALIASES.get(normalized)
        if field and field not in col_map:
            col_map[field] = cell.column
    return col_map


def next_sl_no_for_proforma(teachers, proforma):
    nums = []
    for t in teachers:
        if t.get('proforma') == proforma:
            try:
                nums.append(int(str(t.get('sl_no', '')).strip()))
            except Exception:
                pass
    return str((max(nums) if nums else 0) + 1)


def cell_value(ws, row_idx, col_idx):
    if not col_idx:
        return ''
    val = ws.cell(row=row_idx, column=col_idx).value
    if val is None:
        return ''
    return str(val).strip()


def row_has_any_value(ws, row_idx, max_col):
    for col in range(1, max_col + 1):
        val = ws.cell(row=row_idx, column=col).value
        if val not in (None, ''):
            return True
    return False


def log_activity(action, details=None):
    try:
        logs = []
        if os.path.exists(LOG_PATH):
            with open(LOG_PATH) as f:
                logs = json.load(f)
        entry = {"time": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"), "action": action}
        if details:
            entry["details"] = details
        logs.insert(0, entry)
        with open(LOG_PATH, "w") as f:
            json.dump(logs[:500], f, indent=2)
    except Exception:
        pass


def get_logs(limit=100):
    try:
        with open(LOG_PATH) as f:
            return json.load(f)[:limit]
    except Exception:
        return []


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("teacher_logged_in"):
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("teacher_login"))
        return f(*args, **kwargs)
    return decorated


# ══ Public Routes ══

@app.route("/")
def home():
    return render_template_string(HOME_HTML)


@app.route("/teacher/login", methods=["GET", "POST"])
def teacher_login():
    if request.method == "GET":
        return render_template_string(TEACHER_LOGIN_HTML, schools=get_schools())
    data = request.get_json(silent=True) or request.form
    school = (data.get("school") or "").strip()
    dob    = (data.get("dob") or "").strip()
    if not school or not dob:
        return jsonify({"success": False, "error": "School and date of birth are required."})
    teachers = load_teachers()
    matches  = [t for t in teachers
                if t.get("school") == school and dob_matches(t.get("dob", ""), dob)]
    if not matches:
        return jsonify({"success": False,
                        "error": "No record found. Check your school and date of birth."})
    session["teacher_logged_in"] = True
    session["teacher_school"]    = school
    session["teacher_dob"]       = dob
    log_activity(f"Teacher login: school={school}")
    return jsonify({"success": True, "count": len(matches)})


@app.route("/teacher/logout")
def teacher_logout():
    session.pop("teacher_logged_in", None)
    session.pop("teacher_school", None)
    session.pop("teacher_dob", None)
    return redirect(url_for("teacher_login"))


@app.route("/teacher/dashboard")
@teacher_required
def teacher_dashboard():
    return render_template_string(TEACHER_DASHBOARD_HTML, school=session.get("teacher_school", ""))


@app.route("/api/teacher/mydata")
@teacher_required
def api_teacher_mydata():
    school = session.get("teacher_school")
    dob    = session.get("teacher_dob")
    teachers = load_teachers()
    my_teachers = [t for t in teachers
                   if t.get("school") == school and dob_matches(t.get("dob", ""), dob)]
    return jsonify({"success": True, "teachers": my_teachers,
                    "field_labels": FIELD_LABELS, "field_groups": FIELD_GROUPS})


@app.route("/api/teacher/update", methods=["POST"])
@teacher_required
def api_teacher_update():
    data = request.get_json(silent=True) or {}
    teacher_name = data.get("teacher_name", "").strip()
    proforma     = data.get("proforma", "").strip()
    updates      = data.get("updates", {})
    if not teacher_name or not proforma:
        return jsonify({"success": False, "error": "teacher_name and proforma required"})
    teachers = load_teachers()
    for t in teachers:
        if t.get("teacher_name") == teacher_name and t.get("proforma") == proforma:
            history = t.get("change_history", [])
            changed_fields = []
            for k, v in updates.items():
                if k in PROTECTED_FIELDS:
                    continue
                if t.get(k) != v:
                    history.append({"field": k, "old": t.get(k, ""), "new": v,
                                    "time": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                                    "by": "teacher"})
                    changed_fields.append(k)
                    t[k] = v
            t["change_history"] = history
            if changed_fields:
                t["status"]       = "UPDATED"
                t["last_updated"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                t["updated_by"]   = "teacher"
                log_activity(f"Teacher update: {teacher_name}")
            save_teachers(teachers)
            return jsonify({"success": True})
    return jsonify({"success": False, "error": "Teacher record not found"})


@app.route("/api/teacher/verify", methods=["POST"])
@teacher_required
def api_teacher_verify():
    data = request.get_json(silent=True) or {}
    teacher_name = data.get("teacher_name", "").strip()
    proforma     = data.get("proforma", "").strip()
    teachers = load_teachers()
    for t in teachers:
        if t.get("teacher_name") == teacher_name and t.get("proforma") == proforma:
            t["status"]       = "VERIFIED"
            t["last_updated"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
            t["updated_by"]   = "teacher-verify"
            save_teachers(teachers)
            log_activity(f"Teacher verified: {teacher_name}")
            return jsonify({"success": True})
    return jsonify({"success": False, "error": "Record not found"})


# ══ Admin Routes ══

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "GET":
        return render_template_string(ADMIN_LOGIN_HTML)
    data = request.get_json(silent=True) or request.form
    u = (data.get("username") or "").strip()
    p = (data.get("password") or "").strip()
    if u == ADMIN_USER and p == ADMIN_PASS:
        session["admin_logged_in"] = True
        log_activity(f"Admin login: {u}")
        return jsonify({"success": True})
    return jsonify({"success": False, "error": "Invalid credentials"})


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    return render_template_string(ADMIN_DASHBOARD_HTML)


@app.route("/api/admin/stats")
@admin_required
def api_admin_stats():
    teachers = load_teachers()
    total    = len(teachers)
    verified = sum(1 for t in teachers if t.get("status") == "VERIFIED")
    updated  = sum(1 for t in teachers if t.get("status") == "UPDATED")
    pending  = total - verified - updated
    schools  = len(get_schools())
    duplicate_groups = len(find_duplicate_groups(teachers))
    by_proforma = {}
    for t in teachers:
        pf = t.get("proforma", "OTHER")
        if pf not in by_proforma:
            by_proforma[pf] = {"total": 0, "verified": 0, "updated": 0, "pending": 0}
        by_proforma[pf]["total"] += 1
        st = t.get("status", "PENDING")
        if st == "VERIFIED":   by_proforma[pf]["verified"] += 1
        elif st == "UPDATED":  by_proforma[pf]["updated"]  += 1
        else:                  by_proforma[pf]["pending"]  += 1
    return jsonify({"total": total, "verified": verified, "updated": updated,
                    "pending": pending, "schools": schools, "duplicate_groups": duplicate_groups, "by_proforma": by_proforma})


@app.route("/api/admin/teachers")
@admin_required
def api_admin_teachers():
    teachers = load_teachers()
    q     = request.args.get("q", "").lower()
    pf    = request.args.get("proforma", "")
    st    = request.args.get("status", "")
    limit = int(request.args.get("limit", 9999))
    if q:
        teachers = [t for t in teachers
                    if q in t.get("teacher_name", "").lower() or q in t.get("school", "").lower()]
    if pf:
        teachers = [t for t in teachers if t.get("proforma") == pf]
    if st:
        teachers = [t for t in teachers if t.get("status") == st]
    return jsonify({"teachers": teachers[:limit], "total": len(teachers)})


@app.route("/api/admin/teacher/<path:teacher_name>/<proforma>")
@admin_required
def api_admin_get_teacher(teacher_name, proforma):
    teachers = load_teachers()
    for t in teachers:
        if t.get("teacher_name") == teacher_name and t.get("proforma") == proforma:
            return jsonify({"success": True, "teacher": t,
                            "field_labels": FIELD_LABELS, "field_groups": FIELD_GROUPS})
    return jsonify({"success": False, "error": "Not found"}), 404


@app.route("/api/admin/teacher/update", methods=["POST"])
@admin_required
def api_admin_teacher_update():
    data = request.get_json(silent=True) or {}
    teacher_name = data.get("teacher_name", "").strip()
    proforma     = data.get("proforma", "").strip()
    updates      = data.get("updates", {})
    teachers = load_teachers()
    for t in teachers:
        if t.get("teacher_name") == teacher_name and t.get("proforma") == proforma:
            history = t.get("change_history", [])
            changed = []
            for k, v in updates.items():
                if k in ("change_history", "teacher_name", "proforma"):
                    continue
                if t.get(k) != v:
                    history.append({"field": k, "old": t.get(k, ""), "new": v,
                                    "time": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                                    "by": "admin"})
                    changed.append(k)
                t[k] = v
            t["change_history"] = history
            if changed:
                t["last_updated"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                t["updated_by"]   = "admin"
                log_activity(f"Admin updated: {teacher_name}")
            save_teachers(teachers)
            return jsonify({"success": True})
    return jsonify({"success": False, "error": "Teacher not found"})


@app.route("/api/admin/teacher/reset", methods=["POST"])
@admin_required
def api_admin_teacher_reset():
    data = request.get_json(silent=True) or {}
    teacher_name = data.get("teacher_name", "").strip()
    proforma     = data.get("proforma", "").strip()
    teachers = load_teachers()
    for t in teachers:
        if t.get("teacher_name") == teacher_name and t.get("proforma") == proforma:
            t["status"] = "PENDING"
            t["last_updated"] = ""
            t["updated_by"] = ""
            t["change_history"] = []
            save_teachers(teachers)
            log_activity(f"Admin reset: {teacher_name}")
            return jsonify({"success": True})
    return jsonify({"success": False, "error": "Not found"})


@app.route("/api/admin/teacher/create", methods=["POST"])
@admin_required
def api_admin_teacher_create():
    data = request.get_json(silent=True) or {}
    teacher_name = data.get("teacher_name", "").strip()
    proforma     = data.get("proforma", "").strip()
    school       = data.get("school", "").strip()
    dob          = data.get("dob", "").strip()
    if not teacher_name:
        return jsonify({"success": False, "error": "teacher_name is required"})
    if not proforma:
        return jsonify({"success": False, "error": "proforma is required"})
    if not school:
        return jsonify({"success": False, "error": "school is required"})
    if not dob:
        return jsonify({"success": False, "error": "dob is required"})
    teachers = load_teachers()
    for t in teachers:
        if t.get("teacher_name") == teacher_name and t.get("proforma") == proforma:
            return jsonify({"success": False,
                            "error": f"Teacher '{teacher_name}' in {proforma} already exists."})
    sl_no = data.get("sl_no", "").strip()
    if not sl_no:
        pf_teachers = [t for t in teachers if t.get("proforma") == proforma]
        sl_no = str(len(pf_teachers) + 1)
    new_teacher = {
        "proforma": proforma, "status": "PENDING", "last_updated": "", "updated_by": "",
        "sl_no": sl_no, "district": data.get("district", "Ganjam/Chikiti"),
        "teacher_name": teacher_name, "school": school,
        "general_qualification": data.get("general_qualification", ""),
        "date_general_result": data.get("date_general_result", ""),
        "date_passing": data.get("date_passing", ""),
        "training_qualification": data.get("training_qualification", ""),
        "date_training_result": data.get("date_training_result", ""),
        "date_training_passing": data.get("date_training_passing", ""),
        "otet": data.get("otet", ""), "date_otet": data.get("date_otet", ""),
        "category": data.get("category", ""), "dob": dob,
        "date_first_appt_ss": data.get("date_first_appt_ss", ""),
        "date_regular_teacher": data.get("date_regular_teacher", ""),
        "date_regularisation_consideration": data.get("date_regularisation_consideration", ""),
        "notification_no": data.get("notification_no", ""),
        "rank": data.get("rank", ""),
        "date_joining_levelVB": data.get("date_joining_levelVB", ""),
        "date_joining_levelIII": data.get("date_joining_levelIII", ""),
        "date_joining_levelIV": data.get("date_joining_levelIV", ""),
        "date_superannuation": data.get("date_superannuation", ""),
        "idt_ra": data.get("idt_ra", ""),
        "date_joining_district": data.get("date_joining_district", ""),
        "dp_cp_vigilance": data.get("dp_cp_vigilance", ""),
        "option": data.get("option", ""), "remark": data.get("remark", ""),
        "change_history": [],
    }
    teachers.append(new_teacher)
    save_teachers(teachers)
    schools_list = load_schools()
    if school not in schools_list:
        schools_list.append(school)
        save_schools(schools_list)
    log_activity(f"Admin created teacher: {teacher_name} ({proforma}) @ {school}")
    return jsonify({"success": True, "teacher": new_teacher})


@app.route("/api/admin/teacher/delete", methods=["POST"])
@admin_required
def api_admin_teacher_delete():
    data = request.get_json(silent=True) or {}
    teacher_name = data.get("teacher_name", "").strip()
    proforma     = data.get("proforma", "").strip()
    if not teacher_name or not proforma:
        return jsonify({"success": False, "error": "teacher_name and proforma required"})
    teachers = load_teachers()
    original_count = len(teachers)
    teachers = [t for t in teachers
                if not (t.get("teacher_name") == teacher_name and t.get("proforma") == proforma)]
    if len(teachers) == original_count:
        return jsonify({"success": False, "error": "Teacher not found"})
    save_teachers(teachers)
    log_activity(f"Admin deleted teacher: {teacher_name} ({proforma})")
    return jsonify({"success": True})


# ══ School Management ══

@app.route("/api/admin/schools")
@admin_required
def api_admin_schools():
    return jsonify({"schools": get_schools()})


@app.route("/api/admin/school/add", methods=["POST"])
@admin_required
def api_admin_school_add():
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "error": "School name is required"})
    schools = load_schools()
    if name in schools:
        return jsonify({"success": False, "error": f"School '{name}' already exists"})
    schools.append(name)
    save_schools(schools)
    log_activity(f"Admin added school: {name}")
    return jsonify({"success": True, "message": f"School '{name}' added"})


@app.route("/api/admin/school/rename", methods=["POST"])
@admin_required
def api_admin_school_rename():
    data = request.get_json(silent=True) or {}
    old_name = data.get("old_name", "").strip()
    new_name = data.get("new_name", "").strip()
    if not old_name or not new_name:
        return jsonify({"success": False, "error": "old_name and new_name required"})
    if old_name == new_name:
        return jsonify({"success": True, "message": "No change"})
    schools = load_schools()
    updated_schools = [new_name if s == old_name else s for s in schools]
    if new_name not in updated_schools:
        updated_schools.append(new_name)
    updated_schools = [s for s in updated_schools if s != old_name]
    save_schools(updated_schools)
    teachers = load_teachers()
    count = 0
    for t in teachers:
        if t.get("school") == old_name:
            t["school"] = new_name
            count += 1
    if count:
        save_teachers(teachers)
    log_activity(f"Admin renamed school: '{old_name}' -> '{new_name}' ({count} teachers)")
    return jsonify({"success": True,
                    "message": f"Renamed. {count} teacher record(s) updated."})


@app.route("/api/admin/school/delete", methods=["POST"])
@admin_required
def api_admin_school_delete():
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "error": "School name required"})
    schools = load_schools()
    save_schools([s for s in schools if s != name])
    log_activity(f"Admin deleted school: {name}")
    return jsonify({"success": True, "message": f"School '{name}' removed"})



@app.route("/api/admin/import-excel", methods=["POST"])
@admin_required
def api_admin_import_excel():
    file = request.files.get('file')
    target_proforma = (request.form.get('proforma') or '').strip()
    default_school = (request.form.get('default_school') or '').strip()
    if not file or not getattr(file, 'filename', ''):
        return jsonify({"success": False, "error": "Excel file is required"})
    if not target_proforma:
        return jsonify({"success": False, "error": "Target proforma is required"})
    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xlsm')):
        return jsonify({"success": False, "error": "Only .xlsx or .xlsm files are supported"})
    try:
        file.stream.seek(0)
        wb = openpyxl.load_workbook(file.stream, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read Excel file: {e}"})

    header_row = detect_header_row(ws)
    col_map = build_import_column_map(ws, header_row)
    if 'teacher_name' not in col_map:
        return jsonify({"success": False, "error": "Could not detect a teacher name column in the Excel file"})

    teachers = load_teachers()
    schools = load_schools()
    created = 0
    updated = 0
    skipped = 0
    processed_rows = 0
    warnings = []

    import_fields = [
        'sl_no', 'district', 'teacher_name', 'school', 'general_qualification',
        'date_general_result', 'date_passing', 'training_qualification', 'date_training_result',
        'date_training_passing', 'otet', 'date_otet', 'category', 'dob',
        'date_first_appt_ss', 'date_regular_teacher', 'date_regularisation_consideration',
        'notification_no', 'rank', 'date_joining_levelVB', 'date_joining_levelIII',
        'date_joining_levelIV', 'date_superannuation', 'idt_ra', 'date_joining_district',
        'dp_cp_vigilance', 'option', 'remark', 'status'
    ]

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if not row_has_any_value(ws, row_idx, ws.max_column):
            continue
        processed_rows += 1
        teacher_name = cell_value(ws, row_idx, col_map.get('teacher_name'))
        if not teacher_name:
            skipped += 1
            warnings.append(f'Row {row_idx}: missing teacher name')
            continue
        school = cell_value(ws, row_idx, col_map.get('school')) or default_school
        row_data = {
            'proforma': target_proforma,
            'status': cell_value(ws, row_idx, col_map.get('status')) or 'PENDING',
            'last_updated': '',
            'updated_by': '',
            'sl_no': cell_value(ws, row_idx, col_map.get('sl_no')),
            'district': cell_value(ws, row_idx, col_map.get('district')) or 'Ganjam/Chikiti',
            'teacher_name': teacher_name,
            'school': school,
            'general_qualification': cell_value(ws, row_idx, col_map.get('general_qualification')),
            'date_general_result': cell_value(ws, row_idx, col_map.get('date_general_result')),
            'date_passing': cell_value(ws, row_idx, col_map.get('date_passing')),
            'training_qualification': cell_value(ws, row_idx, col_map.get('training_qualification')),
            'date_training_result': cell_value(ws, row_idx, col_map.get('date_training_result')),
            'date_training_passing': cell_value(ws, row_idx, col_map.get('date_training_passing')),
            'otet': cell_value(ws, row_idx, col_map.get('otet')),
            'date_otet': cell_value(ws, row_idx, col_map.get('date_otet')),
            'category': cell_value(ws, row_idx, col_map.get('category')),
            'dob': cell_value(ws, row_idx, col_map.get('dob')),
            'date_first_appt_ss': cell_value(ws, row_idx, col_map.get('date_first_appt_ss')),
            'date_regular_teacher': cell_value(ws, row_idx, col_map.get('date_regular_teacher')),
            'date_regularisation_consideration': cell_value(ws, row_idx, col_map.get('date_regularisation_consideration')),
            'notification_no': cell_value(ws, row_idx, col_map.get('notification_no')),
            'rank': cell_value(ws, row_idx, col_map.get('rank')),
            'date_joining_levelVB': cell_value(ws, row_idx, col_map.get('date_joining_levelVB')),
            'date_joining_levelIII': cell_value(ws, row_idx, col_map.get('date_joining_levelIII')),
            'date_joining_levelIV': cell_value(ws, row_idx, col_map.get('date_joining_levelIV')),
            'date_superannuation': cell_value(ws, row_idx, col_map.get('date_superannuation')),
            'idt_ra': cell_value(ws, row_idx, col_map.get('idt_ra')),
            'date_joining_district': cell_value(ws, row_idx, col_map.get('date_joining_district')),
            'dp_cp_vigilance': cell_value(ws, row_idx, col_map.get('dp_cp_vigilance')),
            'option': cell_value(ws, row_idx, col_map.get('option')),
            'remark': cell_value(ws, row_idx, col_map.get('remark')),
            'change_history': []
        }
        if not row_data['sl_no']:
            row_data['sl_no'] = next_sl_no_for_proforma(teachers, target_proforma)
        existing = None
        target_key = normalize_teacher_key(teacher_name)
        for t in teachers:
            if normalize_teacher_key(t.get('teacher_name', '')) == target_key and t.get('proforma') == target_proforma:
                existing = t
                break
        if existing:
            changed = False
            for fld in import_fields:
                if fld in ('sl_no', 'teacher_name'):
                    continue
                new_val = row_data.get(fld, '')
                if new_val != '':
                    if existing.get(fld) != new_val:
                        existing[fld] = new_val
                        changed = True
            if row_data.get('school') and row_data['school'] not in schools:
                schools.append(row_data['school'])
            if changed:
                existing['last_updated'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
                existing['updated_by'] = 'admin-import'
                updated += 1
            else:
                skipped += 1
                warnings.append(f'Row {row_idx}: no changes for existing teacher {teacher_name}')
        else:
            if row_data.get('school') and row_data['school'] not in schools:
                schools.append(row_data['school'])
            teachers.append(row_data)
            created += 1

    save_teachers(teachers)
    save_schools(schools)
    log_activity(f'Admin bulk import: proforma={target_proforma}, created={created}, updated={updated}, skipped={skipped}')
    return jsonify({
        'success': True,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'processed_rows': processed_rows,
        'header_row': header_row,
        'mapped_fields': sorted(col_map.keys()),
        'warnings': warnings[:20]
    })


@app.route("/api/admin/duplicates")
@admin_required
def api_admin_duplicates():
    teachers = load_teachers()
    groups = find_duplicate_groups(teachers)
    return jsonify({
        "group_count": len(groups),
        "record_count": sum(len(g.get("records", [])) for g in groups),
        "groups": groups
    })


@app.route("/api/admin/logs")
@admin_required
def api_admin_logs():
    limit = int(request.args.get("limit", 100))
    return jsonify({"logs": get_logs(limit)})


@app.route("/api/admin/export")
@admin_required
def api_admin_export():
    exp_filter = request.args.get("filter", "all")
    exp_status = request.args.get("status", "all")
    teachers = load_teachers()
    if exp_filter and exp_filter != "all":
        teachers = [t for t in teachers if t.get("proforma") == exp_filter]
    if exp_status and exp_status != "all":
        teachers = [t for t in teachers if t.get("status", "PENDING") == exp_status]
    wb = _build_export_excel(teachers)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Chikiti_Teachers_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    log_activity(f"Export: filter={exp_filter}, status={exp_status}, rows={len(teachers)}")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _build_export_excel(teachers):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    PROFORMA_COLORS = {
        "PROFORMA I": "D9E8FF", "PROFORMA II": "D9F0D9", "PROFORMA III": "FFF4CC",
        "PROFORMA IV": "FFE4CC", "PROFORMA V": "F0D9FF", "PROFORMA VI": "FFD9E8",
        "PROFORMA VII": "D9FFFC", "PROFORMA VIII": "F8E1FF",
    }
    STATUS_COLORS = {"VERIFIED": "C6EFCE", "UPDATED": "DDEBF7", "PENDING": "FFEB9C"}
    COLUMNS = list(FIELD_LABELS.keys())
    HEADERS  = [FIELD_LABELS[c] for c in COLUMNS]
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor="0D2366")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin  = Side(style="thin", color="CCCCCC")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    by_proforma = {}
    for t in teachers:
        pf = t.get("proforma", "OTHER")
        by_proforma.setdefault(pf, []).append(t)
    for pf, rows in sorted(by_proforma.items()):
        ws = wb.create_sheet(title=pf[:31])
        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font=header_font; c.fill=header_fill; c.alignment=center_align; c.border=bdr
        ws.row_dimensions[1].height = 32
        for ri, t in enumerate(rows, 2):
            st = t.get("status", "PENDING")
            rc = STATUS_COLORS.get(st, "FFFFFF")
            for ci, col in enumerate(COLUMNS, 1):
                c = ws.cell(row=ri, column=ci, value=str(t.get(col, "")))
                c.fill=PatternFill("solid", fgColor=rc); c.alignment=left_align
                c.border=bdr; c.font=Font(size=9)
            ws.row_dimensions[ri].height = 18
        for ci, col in enumerate(COLUMNS, 1):
            ml = max(len(HEADERS[ci-1]), max((len(str(t.get(col,""))) for t in rows), default=0))
            ws.column_dimensions[get_column_letter(ci)].width = min(ml+4, 40)
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.sheet_view.showGridLines = False
    ws_sum["A1"] = "Chikiti Block — Teacher Verification Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, color="0D2366")
    ws_sum["A2"] = f"Generated: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_sum["A2"].font = Font(size=9, color="888888")
    ws_sum.merge_cells("A1:F1"); ws_sum.merge_cells("A2:F2")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(["Proforma","Total","Verified","Updated","Pending","% Done"], 1):
        c = ws_sum.cell(row=4, column=ci, value=h)
        c.font=Font(bold=True, color="FFFFFF"); c.fill=PatternFill("solid", fgColor="0D2366")
        c.alignment=ca
    ta=va=ua=pa=0
    for ri, (pf, rows) in enumerate(sorted(by_proforma.items()), 5):
        v=sum(1 for t in rows if t.get("status")=="VERIFIED")
        u=sum(1 for t in rows if t.get("status")=="UPDATED")
        p=len(rows)-v-u
        pct=f"{round((v+u)/len(rows)*100)}%" if rows else "0%"
        for ci, val in enumerate([pf,len(rows),v,u,p,pct], 1):
            c=ws_sum.cell(row=ri, column=ci, value=val); c.alignment=ca; c.border=bdr; c.font=Font(size=9)
        ta+=len(rows); va+=v; ua+=u; pa+=p
    lr = 5 + len(by_proforma)
    pct_all = f"{round((va+ua)/ta*100)}%" if ta else "0%"
    for ci, val in enumerate(["TOTAL",ta,va,ua,pa,pct_all], 1):
        c=ws_sum.cell(row=lr, column=ci, value=val)
        c.font=Font(bold=True); c.fill=PatternFill("solid", fgColor="DDEBF7"); c.alignment=ca; c.border=bdr
    for ci, w in enumerate([30,8,10,10,10,8], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    return wb


@app.route("/health")
def health():
    teachers = load_teachers()
    return jsonify({"status": "ok", "teachers": len(teachers),
                    "data_path": DATA_PATH, "data_exists": os.path.exists(DATA_PATH),
                    "time": datetime.datetime.now().strftime("%d/%m/%Y %H:%M")})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
